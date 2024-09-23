import json
import os
import time
from datetime import datetime, timedelta
from functools import partial
from io import BytesIO
from itertools import groupby

import numpy as np
import openpyxl
import pandas as pd
from django.core.exceptions import FieldError
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import transaction
from django.db.models import Prefetch, QuerySet
from django.db.models import Sum, Q, Case, When, FloatField, F
from django.db.models.functions import Abs, Coalesce
from django.http import HttpResponse
from django.http import StreamingHttpResponse
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, mixins, status, serializers
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import AccessToken

from account.handlers.perms import DataFKModel, get_perm_name
from account.handlers.validate_perm import ValidatePermRest
from account.models import User, Perm
from app.logs import app_log
from marketing.order.api.serializers import OrderSerializer, ProductStatisticsSerializer, SeasonalStatisticSerializer, \
    SeasonStatsUserPointSerializer, update_point, update_season_stats_user, update_user_turnover, OrderUpdateSerializer
from marketing.order.models import Order, OrderDetail, SeasonalStatistic, SeasonalStatisticUser
from marketing.price_list.models import SpecialOffer, PriceList, ProductPrice, SpecialOfferProduct
from marketing.product.models import Product
from marketing.sale_statistic.models import UserSaleStatistic
from system_func.models import PeriodSeason, PointOfSeason
from user_system.client_profile.models import ClientProfile
from user_system.employee_profile.models import EmployeeProfile
from utils.constants import maNhomND, so_type, data_status, perm_actions
from utils.datetime_handle import convert_date_format
from utils.model_filter_paginate import filter_data, get_query_parameters, build_absolute_uri_with_params


class GenericApiOrder(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                      mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = OrderSerializer
    queryset = Order.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=Order)]

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return OrderUpdateSerializer
        return OrderSerializer

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['id', 'date_get', 'date_company_get', 'client_id__id'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        serializer.delete(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def users_order(self, request, *args, **kwargs):
        self.serializer_class = self.get_serializer_class()
        user_id = request.query_params.get('user', '')
        if user_id == '':
            user = request.user
        else:
            current_user = request.user
            user = User.objects.filter(id=user_id.upper()).first()
        app_log.info(f"User test: {user}")
        orders = Order.objects.filter(client_id=user)
        response = filter_data(self, request, ['id', 'date_get', 'date_company_get', 'client_id'], queryset=orders,
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ProductStatisticsView(APIView):
    permission_classes = [partial(ValidatePermRest, model=Order)]
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    serializer_class = ProductStatisticsSerializer

    def get(self, request, *args, **kwargs):
        try:
            # Get current user
            user_id = request.query_params.get('user', '')
            user = request.user
            if user_id != '':
                current_user = user
                user = User.objects.filter(id=user_id.upper()).first()
                perm_name = get_perm_name(User)
                if current_user.is_allow(perm_name) or current_user.is_group_has_perm(perm_name):
                    app_log.info(f"User {current_user.id} has permission of {user.id}")
                #     return Response({"error": "User not has permission"}, status=status.HTTP_403_FORBIDDEN)
            app_log.info(f"Test user id: {user}")

            now = datetime.now().date()
            # Set date default
            default_start_date = now - timedelta(days=365)
            default_end_date = now

            # Get params queries
            start_date_1 = request.query_params.get('start_date_1', datetime.strftime(default_start_date, '%d/%m/%Y'))
            end_date_1 = request.query_params.get('end_date_1', datetime.strftime(default_end_date, '%d/%m/%Y'))

            start_date_time = datetime.strptime(start_date_1, '%d/%m/%Y')
            start_date_2 = request.query_params.get('start_date_2',
                                                    datetime.strftime(start_date_time - timedelta(days=365),
                                                                      '%d/%m/%Y'))
            end_date_2 = request.query_params.get('end_date_2',
                                                  datetime.strftime(start_date_time - timedelta(days=1), '%d/%m/%Y'))
            # Get type to calculate
            type_statistic = request.data.get('type_statistic') or request.query_params.get('type_statistic', 'all')
            # Add date period compare to dict
            input_date = {'start_date_1': start_date_1, 'end_date_1': end_date_1,
                          'start_date_2': start_date_2, 'end_date_2': end_date_2}
            # Get products as
            product_query = request.query_params.get('products', '')
            # Combine products to list id
            product_ids = product_query.split(',') if product_query else []

            # Get param variables data
            limit = request.query_params.get('limit', 10)
            page = int(request.query_params.get('page', 1))

            # Get queries and paginate order
            details_1, details_2 = self.paginate_orders(user, input_date, product_ids, type_statistic, limit, page)
            # Process calculate cashback
            statistics = self.statistic_calculate(details_1, details_2)
            # Add data to list
            statistics_list = [{"product_id": k, **v} for k, v in statistics.items()]

            if int(limit) == 0:
                serializer = ProductStatisticsSerializer(statistics_list, many=True)
                response_data = {
                    'data': serializer.data,
                    'total_page': 1,
                    'current_page': 1
                }
                return Response(response_data, status=status.HTTP_200_OK)

            # Paginate data
            paginator = Paginator(statistics_list, limit)
            page_obj = paginator.get_page(page)
            serializer = ProductStatisticsSerializer(page_obj, many=True)

            response_data = {
                'data': serializer.data,
                'total_page': paginator.num_pages,
                # 'total_page': self.total_pages,
                'current_page': page
            }
            if page < paginator.num_pages:
                next_page = request.build_absolute_uri(
                    f'?page={page + 1}&limit={limit}&user={user}&type_statistic={type_statistic}&products={product_query}')
                response_data['next_page'] = next_page
            if page > 1:
                prev_page = request.build_absolute_uri(
                    f'?page={page - 1}&limit={limit}&user={user}&type_statistic={type_statistic}&products={product_query}')
                response_data['prev_page'] = prev_page
            return Response(response_data, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            raise e
            # return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def paginate_orders(self, user, input_date, product_ids, type_statistic, limit, page):
        start_date_1, end_date_1, start_date_2, end_date_2 = self.convert_dates(input_date)
        app_log.info(f"Paginate test page: {page} | {limit}")
        app_log.info(f"Test datetime: {start_date_1} to {end_date_1} | {start_date_2} | to {end_date_2}")
        orders_1 = Order.objects.filter(client_id=user, date_get__gte=start_date_1, date_get__lte=end_date_1)
        orders_2 = Order.objects.filter(client_id=user, date_get__gte=start_date_2, date_get__lte=end_date_2)

        if len(product_ids) > 0:
            orders_1 = orders_1.filter(order_detail__product_id__in=product_ids).distinct()
            orders_2 = orders_2.filter(order_detail__product_id__in=product_ids).distinct()

        match type_statistic:
            case 'special_offer':
                orders_1 = orders_1.filter(Q(new_special_offer__isnull=False) | Q(is_so=True))
                orders_2 = orders_2.filter(Q(new_special_offer__isnull=False) | Q(is_so=True))
            case 'normal':
                orders_1 = orders_1.filter(
                    Q(new_special_offer__isnull=True) & Q(Q(is_so=False) | Q(is_so__isnull=True)))
                orders_2 = orders_2.filter(
                    Q(new_special_offer__isnull=True) & Q(Q(is_so=False) | Q(is_so__isnull=True)))

        orders_1 = orders_1.order_by('-id')
        orders_2 = orders_2.order_by('-id')

        order_ids_1 = orders_1.values_list('id', flat=True)
        order_ids_2 = orders_2.values_list('id', flat=True)

        query_data_1 = Q(order_id__in=order_ids_1)
        query_data_2 = Q(order_id__in=order_ids_2)

        if len(product_ids) > 0:
            query_data_1 &= Q(product_id__in=product_ids)
            query_data_2 &= Q(product_id__in=product_ids)

        details_1 = OrderDetail.objects.filter(query_data_1).values('product_id', 'product_id__name').annotate(
            total_quantity=Sum('order_quantity'),
            total_point=Sum('point_get'),
            total_price=Sum('product_price'),
            total_box=Sum('order_box'),
            total_cashback=Sum(Case(
                When(price_so__isnull=False, then=Coalesce(F('price_so'), 0.0) * F('order_box')),
                default=0,
                output_field=FloatField()
            ))
        )
        details_2 = OrderDetail.objects.filter(query_data_2).values('product_id', 'product_id__name').annotate(
            total_quantity=Sum('order_quantity'),
            total_point=Sum('point_get'),
            total_price=Sum('product_price'),
            total_box=Sum('order_box'),
            total_cashback=Sum(Case(
                When(price_so__isnull=False, then=Coalesce(F('price_so'), 0.0) * F('order_box')),
                default=0,
                output_field=FloatField()
            ))
        )

        return details_1, details_2

    def convert_dates(self, input_date):
        start_date_1 = timezone.make_aware(datetime.strptime(input_date.get('start_date_1'), '%d/%m/%Y'),
                                           timezone.get_current_timezone())
        end_date_1 = timezone.make_aware(
            datetime.strptime(input_date.get('end_date_1'), '%d/%m/%Y') + timedelta(days=1) - timedelta(seconds=1),
            timezone.get_current_timezone())

        start_date_2 = timezone.make_aware(datetime.strptime(input_date.get('start_date_2'), '%d/%m/%Y'),
                                           timezone.get_current_timezone())
        end_date_2 = timezone.make_aware(
            datetime.strptime(input_date.get('end_date_2'), '%d/%m/%Y') + timedelta(days=1) - timedelta(seconds=1),
            timezone.get_current_timezone())

        return start_date_1, end_date_1, start_date_2, end_date_2

    @staticmethod
    def statistic_calculate(details_1, details_2):
        # Combine results into a single dictionary
        combined_results = {}

        for detail in details_1:
            product_id = detail['product_id']
            product_name = detail['product_id__name']

            combined_results[product_id] = {
                "product_name": product_name,
                "current": {
                    "price": detail['total_price'],
                    "point": detail['total_point'],
                    "quantity": detail['total_quantity'],
                    "box": detail['total_box'],
                    "cashback": int(detail['total_cashback'])
                },
                "total_cashback": int(detail['total_cashback'])
            }

        for detail in details_2:
            product_id = detail['product_id']
            product_name = detail['product_id__name']

            if product_id not in combined_results:
                combined_results[product_id] = {
                    "product_name": product_name,
                    "one_year_ago": {}
                }
            # Calculate total cashback for previous period
            first_cashback = combined_results[product_id].get("total_cashback", 0)
            combined_results[product_id]["one_year_ago"] = {
                "price": detail['total_price'],
                "point": detail['total_point'],
                "quantity": detail['total_quantity'],
                "box": detail['total_box'],
                "cashback": int(detail['total_cashback'])
            }
            combined_results[product_id]["total_cashback"] = first_cashback + int(detail['total_cashback'])

        return combined_results


class ExportReport(APIView):
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=Order)]

    def get(self, request):
        start_time = time.time()
        orders = handle_order(request)
        orders = orders.order_by('-date_get', '-id')
        total_items = orders.count()
        limit = request.query_params.get('limit', 20000)
        page = request.query_params.get('page', 1)

        limit = int(limit)
        page = int(page)
        if limit == 0:
            limit = orders.count()
        if orders.count() > 20000:
            limit = 20000
        if page <= 0:
            page = 1
        start_item = (page - 1) * limit
        end_item = page * limit

        if end_item > total_items:
            end_item = total_items - 1

        orders = orders[start_item:end_item]
        start_time2 = time.time()
        workbook = generate_order_excel(orders)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        print(f"Time generate file: {time.time() - start_time2}")

        response = StreamingHttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=BangToa_{datetime.now().strftime("%d-%m-%Y")}.xlsx'

        print(f"Time export to excel: {time.time() - start_time}")
        return response


class TotalStatisticsView(APIView):
    permission_classes = [partial(ValidatePermRest, model=Order)]
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    def get(self, request, *args, **kwargs):
        try:
            # Get current user
            user_id = request.query_params.get('user', '')
            user = request.user
            if user_id != '':
                current_user = user
                user = User.objects.filter(id=user_id.upper()).first()
                perm_name = get_perm_name(User)
                if current_user.is_allow(perm_name) or current_user.is_group_has_perm(perm_name):
                    app_log.info(f"User {current_user.id} has permission of {user.id}")

            now = datetime.now().date()
            default_start_date = now - timedelta(days=365)
            default_end_date = now

            start_date_1 = request.query_params.get('start_date_1',
                                                    datetime.strftime(default_start_date, '%d/%m/%Y'))
            end_date_1 = request.query_params.get('end_date_1', datetime.strftime(default_end_date, '%d/%m/%Y'))

            start_date_time = datetime.strptime(start_date_1, '%d/%m/%Y')
            start_date_2 = request.query_params.get('start_date_2',
                                                    datetime.strftime(start_date_time - timedelta(days=365),
                                                                      '%d/%m/%Y'))
            end_date_2 = request.query_params.get('end_date_2',
                                                  datetime.strftime(start_date_time - timedelta(days=1),
                                                                    '%d/%m/%Y'))

            type_statistic = request.data.get('type_statistic') or request.query_params.get('type_statistic', 'all')

            product_query = request.query_params.get('products', '')
            product_ids = product_query.split(',') if product_query else []

            order_by_field = request.query_params.get('order_by', 'price')

            details_1, details_2 = self.get_order_details(user, start_date_1, end_date_1, start_date_2, end_date_2,
                                                          product_ids, type_statistic)
            statistics, total_statistic = self.calculate_statistics(details_1, details_2)

            sorted_statistics = sorted(statistics.items(),
                                       key=lambda x: x[1].get('current', {}).get(order_by_field, 0),
                                       reverse=True)
            statistics_list = [{"product_id": k, **v} for k, v in sorted_statistics]

            # Get pagination parameters
            limit = int(request.query_params.get('limit', 10))
            page = int(request.query_params.get('page', 1))

            if limit == 0:
                serializer = ProductStatisticsSerializer(statistics_list, many=True)
                response_data = {
                    'total_statistic': total_statistic,
                    'products_statistics': serializer.data,
                    'total_page': 1,
                    'total_count': len(statistics_list),
                    'current_page': 1
                }
                return Response(response_data, status=status.HTTP_200_OK)

            paginator = Paginator(statistics_list, limit)
            page_obj = paginator.get_page(page)
            serializer = ProductStatisticsSerializer(page_obj, many=True)

            response_data = {
                'total_statistic': total_statistic,
                'products_statistics': serializer.data,
                'total_page': paginator.num_pages,
                'total_count': len(statistics_list),
                'current_page': page
            }

            if page < paginator.num_pages:
                next_page = request.build_absolute_uri(
                    f'?page={page + 1}&limit={limit}&user={user_id}&type_statistic={type_statistic}&products={product_query}&start_date_1={start_date_1}&end_date_1={end_date_1}&start_date_2={start_date_2}&end_date_2={end_date_2}&order_by={order_by_field}')
                response_data['next_page'] = next_page
            if page > 1:
                prev_page = request.build_absolute_uri(
                    f'?page={page - 1}&limit={limit}&user={user_id}&type_statistic={type_statistic}&products={product_query}&start_date_1={start_date_1}&end_date_1={end_date_1}&start_date_2={start_date_2}&end_date_2={end_date_2}&order_by={order_by_field}')
                response_data['prev_page'] = prev_page

            return Response(response_data, status=status.HTTP_200_OK)

        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            raise e

    def get_order_details(self, user, start_date_1, end_date_1, start_date_2, end_date_2, product_ids, type_statistic):
        start_date_1, end_date_1, start_date_2, end_date_2 = self.convert_dates(start_date_1, end_date_1, start_date_2,
                                                                                end_date_2)

        orders_1 = Order.objects.filter(client_id=user, date_get__gte=start_date_1, date_get__lte=end_date_1)
        orders_2 = Order.objects.filter(client_id=user, date_get__gte=start_date_2, date_get__lte=end_date_2)

        if product_ids:
            orders_1 = orders_1.filter(order_detail__product_id__in=product_ids).distinct()
            orders_2 = orders_2.filter(order_detail__product_id__in=product_ids).distinct()

        match type_statistic:
            case 'special_offer':
                orders_1 = orders_1.filter(Q(new_special_offer__isnull=False) | Q(is_so=True))
                orders_2 = orders_2.filter(Q(new_special_offer__isnull=False) | Q(is_so=True))
            case 'normal':
                orders_1 = orders_1.filter(
                    Q(new_special_offer__isnull=True) & Q(Q(is_so=False) | Q(is_so__isnull=True)))
                orders_2 = orders_2.filter(
                    Q(new_special_offer__isnull=True) & Q(Q(is_so=False) | Q(is_so__isnull=True)))

        details_1 = orders_1.values('order_detail__product_id', 'order_detail__product_id__name').annotate(
            total_quantity=Sum('order_detail__order_quantity'),
            total_point=Sum('order_detail__point_get'),
            total_price=Sum('order_detail__product_price'),
            total_box=Sum('order_detail__order_box'),
            total_cashback=Sum(Case(
                When(order_detail__price_so__isnull=False,
                     then=Coalesce('order_detail__price_so', 0.0) * F('order_detail__order_box')),
                default=0,
                output_field=FloatField()
            ))
        )

        details_2 = orders_2.values('order_detail__product_id', 'order_detail__product_id__name').annotate(
            total_quantity=Sum('order_detail__order_quantity'),
            total_point=Sum('order_detail__point_get'),
            total_price=Sum('order_detail__product_price'),
            total_box=Sum('order_detail__order_box'),
            total_cashback=Sum(Case(
                When(order_detail__price_so__isnull=False,
                     then=Coalesce('order_detail__price_so', 0.0) * F('order_detail__order_box')),
                default=0,
                output_field=FloatField()
            ))
        )

        return details_1, details_2

    def calculate_statistics(self, details_1, details_2):
        combined_results = {}
        total_price = 0
        total_point = 0
        total_cashback = 0
        total_box = 0
        total_products_type = list()
        for detail in details_1:
            product_id = detail['order_detail__product_id']
            product_name = detail['order_detail__product_id__name']
            total_products_type.append(product_id)
            combined_results[product_id] = {
                "product_name": product_name,
                "current": {
                    "price": detail['total_price'] or 0,
                    "point": detail['total_point'] or 0,
                    "quantity": detail['total_quantity'] or 0,
                    "box": detail['total_box'] or 0,
                    "cashback": int(detail['total_cashback'] or 0)
                },
                "total_cashback": int(detail['total_cashback'] or 0)
            }
            total_price += detail['total_price'] or 0
            total_point += detail['total_point'] or 0
            total_cashback += detail['total_cashback'] or 0
            total_box += detail['total_box'] or 0

        total_price_2 = 0
        total_point_2 = 0
        total_cashback_2 = 0
        total_box_2 = 0
        total_products_type2 = list()
        for detail in details_2:
            product_id = detail['order_detail__product_id']
            product_name = detail['order_detail__product_id__name']
            total_products_type2.append(product_id)

            if product_id not in combined_results:
                combined_results[product_id] = {
                    "product_name": product_name,
                    "current": {
                        "price": 0,
                        "point": 0,
                        "quantity": 0,
                        "box": 0,
                        "cashback": 0
                    },
                    "one_year_ago": {}
                }
            first_cashback = combined_results[product_id].get("total_cashback", 0)
            combined_results[product_id]["one_year_ago"] = {
                "price": detail['total_price'] or 0,
                "point": detail['total_point'] or 0,
                "quantity": detail['total_quantity'] or 0,
                "box": detail['total_box'] or 0,
                "cashback": int(detail['total_cashback'] or 0)
            }
            total_price_2 += detail['total_price'] or 0
            total_point_2 += detail['total_point'] or 0
            total_cashback_2 += detail['total_cashback'] or 0
            total_box_2 += detail['total_box'] or 0
            combined_results[product_id]["total_cashback"] = first_cashback + int(detail['total_cashback'] or 0)

        total_statistic = {
            'current': {
                'total_price': total_price, 'total_point': total_point, 'total_cashback': total_cashback,
                'total_box': total_box, 'type_products': len(total_products_type)
            },
            'last_time': {
                'total_price': total_price_2, 'total_point': total_point_2, 'total_cashback': total_cashback_2,
                'total_box': total_box_2, 'type_products': len(total_products_type2)
            }
        }
        return combined_results, total_statistic

    def convert_dates(self, start_date_1, end_date_1, start_date_2, end_date_2):
        start_date_1 = timezone.make_aware(datetime.strptime(start_date_1, '%d/%m/%Y'), timezone.get_current_timezone())
        end_date_1 = timezone.make_aware(
            datetime.strptime(end_date_1, '%d/%m/%Y') + timedelta(days=1) - timedelta(seconds=1),
            timezone.get_current_timezone())

        start_date_2 = timezone.make_aware(datetime.strptime(start_date_2, '%d/%m/%Y'), timezone.get_current_timezone())
        end_date_2 = timezone.make_aware(
            datetime.strptime(end_date_2, '%d/%m/%Y') + timedelta(days=1) - timedelta(seconds=1),
            timezone.get_current_timezone())

        return start_date_1, end_date_1, start_date_2, end_date_2


class OrderReportView(APIView):
    permission_classes = [partial(ValidatePermRest, model=Order)]
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    def get(self, request):
        start_time = time.time()
        data, num_page, current_page, total_count, page_obj, limit = self.get_query_results(request)
        response_data = {
            'data': data,
            'total_page': num_page,
            'current_page': current_page,
            'total_count': total_count
        }
        try:
            if page_obj.has_next():
                next_page = build_absolute_uri_with_params(request,
                                                           {'page': page_obj.next_page_number(), 'limit': limit})
                response_data['next_page'] = next_page

            if page_obj.has_previous():
                prev_page = build_absolute_uri_with_params(request,
                                                           {'page': page_obj.previous_page_number(), 'limit': limit})
                response_data['prev_page'] = prev_page
        except AttributeError:
            pass

        app_log.info(f'OrderReport2 Query Time: {time.time() - start_time}')
        return Response(response_data, status=status.HTTP_200_OK)

    def get_query_results(self, request, get_type_list=False):
        limit = int(request.data.get('limit') or request.query_params.get('limit', 10))
        page = int(request.data.get('page') or request.query_params.get('page', 1))

        orders = handle_order(request)

        client_ids = orders.values_list('client_id_id', flat=True).distinct()
        client_profiles = {cp.client_id_id: cp for cp in
                           ClientProfile.objects.filter(client_id__in=client_ids)}

        nvtt_ids = {cp.nvtt_id for cp in client_profiles.values() if cp.nvtt_id}
        # client_lv1_ids = {cp.client_lv1_id for cp in client_profiles.values() if cp.client_lv1_id}
        client_lv1_ids = set(orders.values_list('npp_id', flat=True))
        employee_profiles = {ep.employee_id_id: ep for ep in EmployeeProfile.objects.filter(employee_id__in=nvtt_ids)}
        client_lv1_profiles = {cp.client_id_id: cp for cp in ClientProfile.objects.filter(client_id__in=client_lv1_ids)}

        total_orders = orders.count()
        if limit == 0:
            data = [self.get_order_fields(obj, Order, get_type_list, client_profiles, employee_profiles,
                                          client_lv1_profiles) for obj in orders]
            return data, 1, 1, total_orders, None, limit

        paginator = Paginator(orders, limit)
        page_obj = paginator.get_page(page)
        data = [
            self.get_order_fields(obj, Order, get_type_list, client_profiles, employee_profiles, client_lv1_profiles)
            for obj in page_obj]
        return data, paginator.num_pages, page, total_orders, page_obj, limit

    def get_order_fields(self, obj: Order, model, get_type_list, client_profiles, employee_profiles,
                         client_lv1_profiles):
        order_detail = OrderDetail.objects.filter(order_id=obj)
        client_data = client_profiles.get(obj.client_id_id)
        client_name = client_data.register_name if client_data else None
        nvtt_id = obj.nvtt_id if obj.nvtt_id else (client_data.nvtt_id if client_data else None)

        nvtt = employee_profiles.get(nvtt_id, None)
        nvtt_name = nvtt.register_name if nvtt else None
        client_lv1 = client_lv1_profiles.get(obj.npp_id, None)
        client_lv1_name = client_lv1.register_name if client_lv1 else None

        client_info = {
            'id': obj.client_id.id if obj.client_id else '',
            'name': client_name,
            'nvtt': nvtt_name,
            'register_lv1': client_lv1_name
        }
        if get_type_list:
            client_info['list_type'] = obj.list_type

        fields = model._meta.fields
        field_dict = {}
        include_fields = ['id', 'date_get', 'date_company_get', 'date_delay', 'is_so', 'order_point', 'order_price',
                          'note', 'created_at', 'created_by']
        for field in fields:
            field_name = field.name
            if field_name in include_fields:
                include_fields.remove(field_name)
                field_value = getattr(obj, field_name)
                field_dict[field_name] = field_value

        details_fields = OrderDetail._meta.fields
        order_details = []
        for obj in order_detail:
            details_data = {}
            order_details_include = ['product_id', 'order_quantity', 'order_box', 'product_price', 'point_get',
                                     'price_so', 'note']
            for field in details_fields:
                field_name = field.name
                if field_name in order_details_include:
                    field_value = getattr(obj, field_name)
                    if field_name == 'product_id':
                        try:
                            details_data['product_name'] = field_value.name or None
                            field_value = field_value.id
                        except AttributeError:
                            field_value = None
                    details_data[field_name] = field_value
                    order_details_include.remove(field_name)
            order_details.append(details_data)

        field_dict['order_detail'] = order_details
        field_dict['clients'] = client_info
        return field_dict


class ApiSeasonalStatisticUser(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                               mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = SeasonStatsUserPointSerializer
    queryset = SeasonalStatisticUser.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=SeasonalStatisticUser)]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        stats_id = request.query_params.get('stats_id', None)
        if stats_id:
            queryset = queryset.filter(season_stats__id=stats_id)
        response = filter_data(self, request,
                               ['id', 'user__id', 'user__username', 'season_stats__id', 'season_stats__name'],
                               queryset=queryset, **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiSeasonalStatistic(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                           mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = SeasonalStatisticSerializer
    queryset = SeasonalStatistic.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=SeasonalStatistic)]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        stats_type = request.query_params.get('type', None)
        if stats_type:
            queryset = queryset.filter(type=stats_type)

        response = filter_data(self, request, ['id', 'name', 'start_date', 'end_date'],
                               queryset=queryset, **kwargs)
        return Response(response, status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='export')
    def export(self, request, *args, pk=None):
        try:
            season_statistic = self.get_object()
            users_stats = SeasonalStatisticUser.objects.filter(season_stats=season_statistic).values(
                'user__id', 'user__clientprofile__client_lv1_id', 'user__clientprofile__nvtt_id',
                'turn_per_point', 'turn_pick', 'redundant_point', 'total_point'
            )
            df = pd.DataFrame(list(users_stats))

            client_ids = df['user__clientprofile__client_lv1_id']
            nvtt_ids = df['user__clientprofile__nvtt_id']

            client_lv1 = dict(
                ClientProfile.objects.filter(client_id_id__in=client_ids).values_list('client_id', 'register_name'))
            nvtt = dict(
                EmployeeProfile.objects.filter(employee_id_id__in=nvtt_ids).values_list('employee_id', 'register_name'))

            df.rename(columns={
                'user__id': 'Mã Khách Hàng',
                'user__clientprofile__client_lv1_id': 'NPP',
                'user__clientprofile__nvtt_id': 'NVTT',
                'turn_per_point': 'Điểm/Tem',
                'turn_pick': 'Số tem',
                'redundant_point': 'Điểm dư',
                'total_point': 'Tổng điểm'
            }, inplace=True)
            df['NPP'] = df['NPP'].map(client_lv1)
            df['NVTT'] = df['NVTT'].map(nvtt)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
            response['Content-Disposition'] = f'attachment; filename="{season_statistic.name}.xlsx"'
            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Season Stats')
                worksheet = writer.sheets['Season Stats']
                # Set column widths (Excel column width units)
                worksheet.column_dimensions['A'].width = 124 / 7  # Approximation to Excel units
                worksheet.column_dimensions['B'].width = 140 / 7
                worksheet.column_dimensions['C'].width = 140 / 7

            return response
        except Exception as e:
            # return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
            raise e


class OrderSOCount(APIView):
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=SeasonalStatistic)]

    def get(self, request):
        today = datetime.today().date()
        user_id = request.query_params.get('user', None)
        so_id = request.query_params.get('so_id', None)
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        get_date = request.query_params.get('get_date', f'{today}')

        try:
            user = request.user
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return Response({'message': f'not found user with id {user_id}'})
            print(f"Test user: {user}")
            base_filter = Q(client_id=user) & Q(new_special_offer__isnull=False)

            if from_date:
                from_date = datetime.strptime(from_date, '%Y-%m-%d')
            if to_date:
                to_date = datetime.strptime(to_date, '%Y-%m-%d')
            if from_date and to_date:
                base_filter &= Q(date_get__gte=from_date, date_get__lte=to_date)
            elif get_date == 'all':
                pass
            else:
                try:
                    get_date = datetime.strptime(get_date, '%Y-%m-%d')
                except ValueError:
                    return Response({'message': "get_date phải có định dạng là 'YYYY-MM-DD' hoặc 'all'"})
                base_filter &= Q(date_get=get_date)
            if so_id:
                base_filter &= Q(new_special_offer__id=so_id)
            get_so_order = (Order.objects.filter(base_filter)
                            # .exclude(Q(new_special_offer__type_list=so_type.consider_user))
                            )
            order_ids_used = get_so_order.values_list('id', flat=True).distinct()
            so_ids_used = get_so_order.values_list('new_special_offer__id', flat=True).distinct()

            so_objs = SpecialOffer.objects.filter(id__in=list(so_ids_used),
                                                  orders__id__in=list(order_ids_used)).distinct()

            data = self.serializer(so_objs, get_so_order)

            response = {
                'user': {
                    'id': user.id,
                    'register_name': user.clientprofile.register_name
                },
                'data': data,
            }

            return Response(response)
        except Exception as e:
            raise e

    def serializer(self, so_objs, get_so_order):
        data = []
        for so in so_objs:
            orders = get_so_order.filter(new_special_offer=so)
            order_ids = orders.values_list('id', flat=True)
            total_box = OrderDetail.objects.filter(order_id__in=order_ids).aggregate(
                total_box=Sum('order_box'))['total_box'] or 0

            input_data = {
                'so_id': so.id,
                'time_used': orders.count(),
                'total_used_box': total_box,
                'orders_used': list(orders.values_list('id', flat=True)),
            }
            data.append(input_data)
        return data


def handle_order(request) -> QuerySet[Order]:
    query, strict_mode, limit, page, order_by, from_date, to_date, date_field = get_query_parameters(request)
    app_log.info(f"Params: \n__query: {query}"
                 f"\n__strict_mode: {strict_mode}"
                 f"\n__limit: {limit}"
                 f"\n__page: {page}"
                 f"\n__order_by: {order_by}"
                 f"\n__from_date: {from_date}"
                 f"\n__to_date: {to_date}"
                 f"\n__date_field: {date_field}"
                 f"")
    nvtt_query = request.query_params.get('nvtt', '')
    npp_query = request.query_params.get('npp', '')
    daily_query = request.query_params.get('daily', '')

    order_by = '-date_get' if order_by == '' else order_by

    orders = Order.objects.all()

    if from_date or to_date:
        try:
            if from_date:
                from_date = datetime.strptime(from_date, '%d/%m/%Y')
                orders = orders.filter(**{f'{date_field}__gte': from_date})
            if to_date:
                to_date = datetime.strptime(to_date, '%d/%m/%Y') + timedelta(days=1) - timedelta(seconds=1)
                orders = orders.filter(**{f'{date_field}__lte': to_date})
        except ValueError:
            pass

    if nvtt_query and nvtt_query != '':
        nvtt_list = nvtt_query.split(',')
        app_log.info(f'Test nvtt list: {nvtt_list}')
        nvtt_ids = EmployeeProfile.objects.filter(
            Q(register_name__in=nvtt_list) | Q(employee_id__in=nvtt_list)
        ).values_list('employee_id', flat=True)
        client_profile_query = Q(client_id__clientprofile__nvtt_id__in=nvtt_ids)
        orders = orders.filter(client_profile_query)

    if npp_query and npp_query != '':
        npp_list = npp_query.split(',')
        app_log.info(f'Test npp list: {npp_list}')
        npp_ids = ClientProfile.objects.filter(
            Q(register_name__in=npp_list, is_npp=True) | Q(client_id__in=npp_list, is_npp=True)
        ).values_list('client_id', flat=True)
        client_profile_query = Q(client_id__in=npp_ids)
        orders = orders.filter(client_profile_query)

    if daily_query and daily_query != '':
        daily_list = daily_query.split(',')
        exclude_groups = ["NVTT", "TEST", maNhomND]

        daily_ids = ClientProfile.objects.filter(
            Q(register_name__in=daily_list) | Q(client_id__in=daily_list)
        ).exclude(client_group_id__name__in=exclude_groups, is_npp=True).values_list('client_id', flat=True)

        client_query = Q(client_id__in=daily_ids)
        orders = orders.filter(client_query)

    if query != '':
        query_parts = query.split(',')
        order_query = Q()
        order_detail_query = Q()
        client_profile_query = Q()
        for part in query_parts:
            order_query |= Q(id__icontains=part) | Q(date_get__icontains=part) | Q(date_company_get__icontains=part)
            order_detail_query |= Q(order_detail__product_id__id__icontains=part) | Q(
                order_detail__product_id__name__icontains=part)
            client_profile_query |= Q(client_id__clientprofile__register_name__icontains=part) | Q(
                client_id__clientprofile__nvtt_id__icontains=part) | Q(
                client_id__clientprofile__client_group_id__id__icontains=part)

        if strict_mode:
            orders = orders.filter(order_query & order_detail_query & client_profile_query)
        else:
            orders = orders.filter(order_query | order_detail_query | client_profile_query)

    # valid_fields = [f.name for f in Order._meta.get_fields()]
    # try:
    #     if order_by in valid_fields or order_by.lstrip('-') in valid_fields:
    #         orders = orders.order_by(order_by)
    #     else:
    #         orders = orders.order_by('created_at')
    # except FieldError:
    #     orders = orders.order_by('id')

    order_by_fields = order_by.split(',')
    normalized_fields = []
    for field in order_by_fields:
        normalized_field = field.strip().lstrip('-')
        if normalized_field in [f.name for f in Order._meta.get_fields()]:
            normalized_fields.append(field.strip())  # Maintain order direction

    if normalized_fields:
        orders = orders.order_by(*normalized_fields)
    else:
        try:
            orders = orders.order_by('created_at')
        except FieldError:
            orders = orders.order_by('id')
        except Exception as e:
            app_log.error("Error in order_by with default fields")
            raise e

    orders = orders.select_related('client_id').prefetch_related(
        Prefetch('order_detail', queryset=OrderDetail.objects.select_related('product_id'))
    ).distinct()
    # specific = orders.filter(id='MTN240901256').first()
    # app_log.info(f"Test EXPORT: {specific}")
    return orders


def generate_order_excel(orders: list[Order]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    title_font = Font(name='Times New Roman', bold=True, size=20)
    note_font = Font(name='Times New Roman', size=11)
    header_font = Font(name='Times New Roman', bold=True, color='FFFFFF')
    date_font = Font(name='Times New Roman', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='33cc33', end_color='33cc33', fill_type='solid')
    bold_font = Font(name='Times New Roman', bold=True)

    border_style = Side(style='medium')
    full_border_style = Border(left=border_style, right=border_style, top=border_style, bottom=border_style,
                               diagonal=border_style, diagonal_direction=0)

    sheet.merge_cells('A1:N1')
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Bảng thống kê toa thuốc {datetime.now().strftime("%d-%m-%Y")}'
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    total_price = orders.aggregate(total_price=Sum('order_price'))['total_price']

    total_price = OrderDetail.objects.filter(order_id__in=orders).aggregate(
        total_price=Sum('product_price'))['total_price']
    sheet.merge_cells('A2:N2')
    note_cell = sheet.cell(row=2, column=1)
    note = f'Ngày thống kê: {datetime.now().strftime("%d/%m/%Y")}   ||   Tổng doanh thu: {total_price}   ||   Số lượng bản kê: {orders.count()}'

    note_cell.value = note
    note_cell.font = note_font
    note_cell.alignment = center_alignment

    columns = [
        'Mã toa',
        'Loại bảng kê', 'Mã khách hàng', 'Tên Khách hàng', 'Khách hàng cấp 1', 'NVTT',
        'Ngày nhận toa', 'Người tạo toa', 'Ngày nhận hàng', 'Ngày gửi trễ', 'Ghi chú',
        'Mã sản phẩm', 'Tên sản phẩm', 'Số lượng', 'Số thùng', 'Đơn giá', 'Thành tiền',
        'Đơn giá KM', 'Điểm đạt'
    ]

    column_widths = {
        'Mã toa': 106,
        'Loại bảng kê': 92,
        'Mã khách hàng': 98,
        'Tên Khách hàng': 136,
        'Khách hàng cấp 1': 140,
        'NVTT': 144,
        'Ngày nhận toa': 92,
        'Người tạo toa': 86,
        'Ngày nhận hàng': 98,
        'Ngày gửi trễ': 84,
        'Ghi chú': 72,
        'Mã sản phẩm': 84,
        'Tên sản phẩm': 196,
        'Số lượng': 64,
        'Số thùng': 68,
        'Đơn giá': 80,
        'Thành tiền': 82,
        'Đơn giá KM': 80,
        'Điểm đạt': 64
    }

    for col_num, column_title in enumerate(columns, 1):
        cell = sheet.cell(row=4, column=col_num)  # Thêm tiêu đề cột từ dòng 4
        cell.value = column_title
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = full_border_style

        col_letter = get_column_letter(col_num)
        if column_title in column_widths:
            sheet.column_dimensions[col_letter].width = column_widths[column_title] / 7.2
        else:
            sheet.column_dimensions[col_letter].width = 120 / 7.2

    row_num = 5

    client_profiles = {cp.client_id_id: cp for cp in
                       ClientProfile.objects.filter(client_id__in=[o.client_id_id for o in orders])}
    # client_lv1_ids = [client_profiles[o.client_id_id].client_lv1_id for o in orders if
    #                   o.client_id_id in client_profiles and client_profiles[o.client_id_id].client_lv1_id]
    npp_ids = orders.values_list('npp_id', flat=True)
    npp_ids = list(set(npp_ids))
    npp_profiles = {cp.client_id_id: cp for cp in ClientProfile.objects.filter(client_id__in=npp_ids)}

    # price_lists = orders.values_list('price_list_id_id', flat=True).distinct()

    nvtt_ids = set()
    for o in orders:
        if o.nvtt_id and o.nvtt_id not in nvtt_ids:
            nvtt_ids.add(o.nvtt_id)
        elif o.client_id_id in client_profiles and client_profiles[o.client_id_id].nvtt_id:
            nvtt_ids.add(client_profiles[o.client_id_id].nvtt_id)

    employee_profiles = {ep.employee_id_id: ep for ep in EmployeeProfile.objects.filter(employee_id__in=nvtt_ids)}
    app_log.info(f"__ Count order: {orders.count()}")
    for i, order in enumerate(orders):
        if order.id == 'MTN240901256':
            app_log.info(f"Test EXPORT 1: {order}")
        # Get client data from query result
        client_data = client_profiles.get(order.client_id_id)
        # Handle change datetime format date_company_get
        try:
            date_obj_local = order.date_company_get.astimezone()
            date_send = datetime.strftime(date_obj_local, "%d/%m/%Y")
        except Exception as e:
            date_send = ''

        # Handle change date format date_get
        try:
            date_get = datetime.strftime(order.date_get, "%d/%m/%Y")
        except TypeError:
            date_get = ''

        # Get type list
        type_list = order.list_type if order.list_type and order.list_type != '' else 'cấp 2 gửi'

        # Get client lv1/npp name
        npp_name = ''
        if order.npp_id not in [None, '']:
            npp_name = npp_profiles[order.npp_id].register_name
        # Get nvtt name
        nvtt_name = ''
        nvtt_id = order.nvtt_id if order.nvtt_id else (client_data.nvtt_id if client_data else None)
        if nvtt_id and nvtt_id in employee_profiles:
            nvtt_name = employee_profiles[nvtt_id].register_name
        # Generate note as dict for json decode
        note_dict = {}
        if order.note != '':
            try:
                note_dict = json.loads(order.note)
            except (json.JSONDecodeError, TypeError):
                pass
        # Trying get dict[key] from notes dict
        noting = note_dict.get('notes', '')

        # Create data for a row
        data_list = [
            order.id,
            type_list,
            client_data.client_id_id if client_data else '',  # Mã khách hàng
            client_data.register_name if client_data else '',  # Tên Khách hàng
            npp_name,  # Khách hàng cấp 1
            nvtt_name,  # NVTT
            date_send,
            order.created_by,
            date_get,
            order.date_delay if order.date_delay else 0,
            noting,
        ]
        if order.order_detail.all().count() > 0:
            for detail in order.order_detail.all():
                product_id = ''
                product_name = ''
                if detail.product_id:
                    product_id = detail.product_id.id
                    product_name = detail.product_id.name
                total_price = detail.product_price or 0
                price_so = detail.price_so if detail.price_so else ''

                # Get price from note
                # try:
                #     note_price = json.loads(detail.note).get('price')
                #     price = float(note_price) if note_price else None
                # except (json.JSONDecodeError, TypeError, ValueError):
                #     price = None
                # if price is None:
                #     if order.new_special_offer_id:
                #         price = get_special_offer_price(detail.product_id, order.new_special_offer_id)
                #     elif order.price_list_id_id:
                #         price = get_product_price(detail.product_id, order.price_list_id_id)
                try:
                    price = total_price / detail.order_quantity
                except ZeroDivisionError:
                    price = 0
                details_data = [
                    product_id,
                    product_name,
                    detail.order_quantity,
                    detail.order_box,
                    price,
                    detail.product_price,
                    price_so,
                    detail.point_get
                ]
                export_data = data_list + details_data
                sheet.append(export_data)

                for col, value in enumerate(export_data, 1):
                    cell = sheet.cell(row=row_num, column=col)
                    cell.font = note_font
                    if col in {7, 9}:
                        cell.alignment = center_alignment
                        cell.font = date_font
                    if col == 18:  # 'Đơn giá KM' column
                        cell = sheet.cell(row=row_num, column=col)
                        cell.font = bold_font

                row_num += 1
        else:
            details_data = [
                '',
                '',
                0,
                0,
                order.order_price,
                0,
                0,
                0
            ]
            export_data = data_list + details_data
            sheet.append(export_data)

    return workbook


def get_product_price(product_id, price_list_id):
    product_price = ProductPrice.objects.filter(
        product_id=product_id, price_list_id=price_list_id
    ).first()
    return product_price.price if product_price else None


def get_special_offer_price(product_id, special_offer_id):
    special_offer_product = SpecialOfferProduct.objects.filter(
        product_id=product_id, special_offer_id=special_offer_id
    ).first()
    return special_offer_product.price if special_offer_product else None


class ApiImportOrder(APIView):
    def post(self, request):
        file = request.FILES.get('file_import', None)
        if not file:
            return Response({'message': f'file_import is required'})
        file_extension = os.path.splitext(file.name)[1].lower()

        if file_extension not in ['.xlsx']:
            return Response({'message': 'File must be .xlsx'}, status=status.HTTP_400_BAD_REQUEST)
        # data = get_excel_to_dict(file)
        # print(data)
        # return Response({'data': data})
        try:
            data = get_excel_to_dict(file)

            success, error = create_order(data)
            app_log.info(f"Import success: \n{success}")
            app_log.info(f"Import error: \n{error}")
            return Response({
                'message': 'ok',
                'success': success,
                'errors': error
            }, status=status.HTTP_200_OK)
        except Exception as e:
            # return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            raise e


def create_order(data):
    error_data = list()
    update_list = list()

    for orders_data in data:
        data_lines = orders_data['lines_number']
        try:
            with transaction.atomic():
                client_id = orders_data.get('client_id')
                # Trying query get user by client id
                client: User = User.objects.filter(id=client_id).first()
                # When client not exist
                if not client:
                    # Print data to errors
                    raise ValueError(f'user {orders_data.get("client_id")} không tồn tại')
                # Handle date
                min_date = orders_data.get('min_date_get')
                max_date = orders_data.get('max_date_get')

                # Get available price list
                price_lists = PriceList.objects.filter(
                    Q(date_start__lte=min_date, date_end__gte=min_date) |
                    Q(date_start__lte=max_date, date_end__gte=max_date)
                )
                price_lists_id = price_lists.values_list('id', flat=True).distinct()

                available_perm = [f'{perm_actions["create"]}_{get_perm_name(PriceList)}_{pl_id}' for pl_id in
                                  price_lists_id]
                all_client_perm = client.get_all_user_perms()
                all_prl_perm = all_client_perm.filter(name__in=available_perm)
                prl_ids = list()
                if all_prl_perm.exists():
                    prl_ids = all_prl_perm.values_list('object_id', flat=True).distinct()
                pl = price_lists.filter(id__in=prl_ids)
                app_log.info(f"PL of user {client_id}: {pl.count()} - {pl.first()}")
                for order in orders_data.get('orders', []):
                    # Split detail data
                    details_data = order.pop('order_details')
                    # Get client id
                    nvtt = User.objects.filter(clientprofile__register_name__icontains=order['nvtt'],
                                               group_user__name='nvtt')
                    nvtt_id = ''
                    if nvtt.exists():
                        nvtt_id = nvtt.first().id
                    # Create Order data
                    order_data = Order(
                        client_id=client,
                        list_type=order['type_list'],
                        date_get=order['date_get'],
                        date_company_get=order['date_company_get'],
                        date_delay=order['date_delay'],
                        nvtt_id=nvtt_id,
                        created_by=order['created_by'],
                        status=data_status.active
                    )
                    noting = order['note']
                    # Jsonify note
                    note = json.dumps({'notes': noting})
                    order_data.note = note

                    # Call save() to create Order
                    order_data.save()

                    detail_order = list()
                    turnover_minus = order.get('minus_turnover', None)
                    count_so = order.get('count_so', None)
                    count_turnover = False
                    if count_so not in ['', 'nan', None]:
                        count_turnover = True
                    so_data = {
                        'is_so': False,
                        'minus': turnover_minus,
                        'count': count_turnover
                    }
                    total_point = 0
                    total_price = 0
                    detail_error = False
                    success_line = []
                    # Loop to details_data
                    for detail in details_data:
                        if detail_error:
                            break
                        # Trying get product object
                        product = Product.objects.filter(id=detail['product_id']).first()
                        point = detail.get('point', None)
                        # If product not found
                        if not product:
                            raise ValueError(f'quy cách {detail["product_id"]} không tồn tại')

                        if point is None:
                            if pl.count() > 1:
                                raise ValueError(f'user có {pl.count()} bảng giá, không lấy được điểm cho sản phẩm')

                            elif pl.count() == 1:
                                product_price = ProductPrice.objects.filter(price_list=pl.first(), product=product)
                                if product_price.exists():
                                    product_price = product_price.first()
                                    point = product_price.point
                                    point = point if point else 0
                                else:
                                    point = 0
                            else:
                                point = 0

                        note = {
                            'id': product.id,
                            'price': detail['price'],
                            'point': point,
                            'to_money': ''
                        }
                        note = json.dumps(note)
                        points = point * detail['box']
                        price = detail['total_price'] if detail['total_price'] else 0
                        price_so = detail['price_so'] if detail['price_so'] not in ['', None, 'nan'] else None
                        order_detail = OrderDetail(
                            order_id=order_data,
                            product_id=product,
                            order_quantity=detail['quantity'],
                            order_box=detail['box'],

                            note=note,
                            product_price=price,
                            point_get=points,
                            price_so=price_so
                        )
                        total_price += price
                        total_point += points
                        #     total_price += detail['price']
                        detail_order.append(order_detail)
                        if price_so and not so_data.get('is_so'):
                            so_data['is_so'] = True
                        success_line.append(detail.get('line_number'))
                    if detail_error:
                        order.delete()
                        continue
                    order_data.order_point = total_point
                    order_data.order_price = total_price
                    if so_data.get('is_so'):
                        order_data.is_so = so_data.get('is_so')
                    order_data.save()
                    # if has_nvtt and all(product_prices):
                    # handle_after_order(client, total_point, total_price)
                    update_list.append(
                        {
                            'success_line': success_line,
                            'group_order': order.get('group_order'),
                            'client_id': orders_data.get('client_id'),
                            'order_id': order_data.id,
                        }
                    )
                    data_lines = [number for number in data_lines if number not in success_line]
                    OrderDetail.objects.bulk_create(detail_order)
                    update_point(client)
                    update_season_stats_user(client, order_data.date_get)
                    is_so = order_data.is_so if order_data.is_so in [False, True] else False
                    update_user_turnover(client, order_data, is_so, so_data=so_data)
        except ValueError as e:
            error = str(e)
            detail_message = error.split('\\n')[1]
            error_ = {
                'line': data_lines,
                'message': f"{detail_message}"
            }
            error_data.append(error_)
        except Exception as e:
            data_error = {
                'error_lines': data_lines,
                'message': f'lỗi import: {e}'
            }
            error_data.append(data_error)

    return update_list, error_data


def get_excel_to_dict(file):
    # Đọc dữ liệu từ file Excel
    df = pd.read_excel(file, engine='openpyxl')

    # Định nghĩa mapping giữa tên cột Excel và key trong dict
    column_mapping = {
        'Mã toa': 'group_order',
        'Loại bảng kê': 'type_list',
        'Mã khách hàng': 'client_id',
        'Tên Khách hàng': 'client_name',
        'Khách hàng cấp 1': 'client_lv1',
        'NVTT': 'nvtt',
        'Ngày nhận toa': 'date_company_get',
        'Người tạo toa': 'created_by',
        'Ngày nhận hàng': 'date_get',
        'Ngày gửi trễ': 'date_delay',
        'Ghi chú': 'note',
        'Mã sản phẩm': 'product_id',
        'Tên sản phẩm': 'product_name',
        'Số lượng': 'quantity',
        'Số thùng': 'box',
        'Đơn giá': 'price',
        'Thành tiền': 'total_price',
        'Đơn giá KM': 'price_so',
        'Điểm 1 thùng': 'point',
        'Tính doanh số': 'count_so',
        'Trừ doanh số': 'minus_turnover'
    }

    # Áp dụng việc đổi tên cột
    df.rename(columns=column_mapping, inplace=True)
    df['line_number'] = df.index + 1
    # Xử lý các giá trị NaN, +inf, -inf
    df.replace({np.inf: None, -np.inf: None, np.nan: None}, inplace=True)

    df['date_company_get'] = df['date_company_get'].apply(convert_date_format)
    df['date_get'] = df['date_get'].apply(convert_date_format)

    # Tính toán ngày nhỏ nhất và lớn nhất cho mỗi client_id
    client_dates = df.groupby('client_id')['date_get'].agg(min_date_get='min', max_date_get='max').reset_index()

    # Gộp ngày vào df
    df = df.merge(client_dates, on='client_id')

    # Sắp xếp dữ liệu
    df.sort_values(by=['client_id', 'date_get'], inplace=True)

    grouped = df.groupby(['client_id', 'group_order'])
    result = []
    for (client_id, group_order), group in grouped:
        order_data = {
            'group_order': group_order,
            'type_list': group['type_list'].iloc[0],
            'date_company_get': group['date_company_get'].iloc[0],
            'created_by': group['created_by'].iloc[0],
            'date_get': group['date_get'].iloc[0],
            'date_delay': group['date_delay'].iloc[0],
            'client_lv1': group['client_lv1'],
            'nvtt': group['nvtt'],
            'note': group['note'].iloc[0],

            'order_details': group[[
                'product_id', 'product_name', 'quantity', 'box', 'price',
                'total_price', 'price_so', 'point', 'count_so', 'minus_turnover', 'line_number'
            ]].to_dict(orient='records')
        }
        result.append((client_id, order_data))

    # Re-group by client_id to nest orders within clients
    final_result = []
    result.sort(key=lambda x: x[0])  # Ensure data is sorted by client_id before regrouping
    for key, group in groupby(result, key=lambda x: x[0]):
        client_orders = [item[1] for item in group]
        client_info = df[df['client_id'] == key].iloc[0]
        lines_number = [ln for order in client_orders for ln in
                        [detail['line_number'] for detail in order['order_details']]]
        client_group = {
            'client_id': key,
            'client_name': client_info['client_name'],
            'min_date_get': client_info['min_date_get'],
            'max_date_get': client_info['max_date_get'],
            'lines_number': lines_number,
            'orders': client_orders
        }
        final_result.append(client_group)

    return final_result
