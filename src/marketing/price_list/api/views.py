import math
import os
from functools import partial

import numpy as np
import openpyxl
import pandas
import pandas as pd
from django.db import transaction
from django.db.models import F
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, mixins, status
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication

from account.handlers.perms import perm_queryset, get_perm_name, export_users_has_perm
from account.handlers.restrict_serializer import add_perm, create_full_perm, list_user_has_perm
from account.handlers.validate_perm import ValidatePermRest
from account.models import User
from app.logs import app_log
from marketing.price_list.api.serializers import PriceListSerializer, SpecialOfferSerializer, PriceList2Serializer, \
    SpecialOfferProductSerializer
from marketing.price_list.models import PriceList, SpecialOffer, ProductPrice, SpecialOfferProduct
from marketing.product.models import Product
from utils.constants import so_type, data_status, so_type_list, perm_actions
from utils.datetime_handle import convert_date_format
from utils.model_filter_paginate import filter_data


class GenericApiPriceList(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                          mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = PriceListSerializer
    # queryset = PriceList.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=PriceList)]

    def get_queryset(self):
        app_log.info(f"Getting query set")
        user = self.request.user
        return perm_queryset(self, user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        user_id = request.query_params.get('user', None)
        if user_id:
            user = User.objects.filter(id=user_id)
            if not user.exists():
                return Response({'message': f'user id {user_id} not found'})
            user = user.first()
            queryset = perm_queryset(self, user)
        response = filter_data(self, request, ['id', 'name', 'date_start', 'date_end'],
                               queryset=queryset, **kwargs)
        return Response(response, status.HTTP_200_OK)

    @action(methods=['get'], detail=False, url_path='now', url_name='now')
    def now(self, request, *args, **kwargs):
        today = timezone.localdate()
        queryset = self.get_queryset()
        user_id = request.query_params.get('user', None)
        if user_id:
            user = User.objects.filter(id=user_id)
            if not user.exists():
                return Response({'message': f'user id {user_id} not found'})
            user = user.first()
            queryset = perm_queryset(self, user)
        order_by = request.data.get('order_by', '') or request.query_params.get('order_by', '')

        queryset = queryset.filter(date_start__lte=today, date_end__gte=today).exclude(
            status='deactivate'
        ).order_by('created_at')
        response = filter_data(self, request, ['id', 'name', 'date_start', 'date_end'], queryset=queryset,
                               **kwargs)
        return Response(response)

    def export_products(self, request, *args, **kwargs):
        try:
            pk = kwargs.get('pk')
            pl: PriceList = PriceList.objects.filter(id=pk).first()
            if not pl:
                return Response({'message': f'không tìm thấy bảng giá với id {pk}'}, status=404)

            products = ProductPrice.objects.filter(price_list=pl)

            # Tạo Workbook mới
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Ghi tiêu đề cột
            columns = ['Mã thuốc', 'Tên thuốc', 'Số lượng', 'Đơn giá', 'Điểm']
            for col_num, column_title in enumerate(columns, 1):
                column_letter = get_column_letter(col_num)
                sheet[f'{column_letter}1'] = column_title

            # Ghi dữ liệu sản phẩm vào các hàng tiếp theo
            for row_num, product_price in enumerate(products, 2):
                sheet[f'A{row_num}'] = product_price.product.id  # Cột Mã thuốc
                sheet[f'B{row_num}'] = product_price.product.name  # Cột Tên thuốc
                sheet[f'C{row_num}'] = product_price.quantity_in_box  # Cột Số lượng
                sheet[f'D{row_num}'] = product_price.price  # Cột Đơn giá
                sheet[f'E{row_num}'] = product_price.point if product_price.point is not None else 0  # Cột Điểm

            # Tạo HTTP response với nội dung file Excel
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename=product_prices_{pk}.xlsx'

            # Lưu Workbook vào response
            workbook.save(response)

            return response

        except Exception as e:
            app_log.error(f"Error in price_list export_products: {e}")
            return Response({'message': f'Error occurred: {str(e)}'}, status=500)

    def export_users(self, request, *args, **kwargs):
        try:
            pk = kwargs.get('pk')
            pl: PriceList = PriceList.objects.filter(id=pk).first()
            if not pl:
                return Response({'message': f'không tìm thấy bảng giá với id {pk}'}, status=404)
            perm_name = get_perm_name(pl)
            perm_name_pk = perm_name + f'_{pk}'

            workbook = export_users_has_perm(pl, pk)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="UserDungBangGia_{pk}.xlsx"'

            # Lưu workbook vào response
            workbook.save(response)

            return response

        except Exception as e:
            app_log.error(f"Error in price_list export_users: {e}")
            # return Response({'message': f'Error occurred: {str(e)}'}, status=500)
            raise e


class ApiSpecialOffer(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                      mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = SpecialOfferSerializer
    # queryset = SpecialOffer.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=PriceList)]

    def get_queryset(self):
        user = self.request.user
        return perm_queryset(self, user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        user_id = request.query_params.get('user', None)
        if user_id:
            user = User.objects.filter(id=user_id)
            if not user.exists():
                return Response({'message': f'user id {user_id} not found'})
            user = user.first()
            queryset = perm_queryset(self, user)
        response = filter_data(self, request, ['id', 'name', 'status', 'type_list', 'priority'],
                               queryset=queryset, **kwargs)
        return Response(response, status.HTTP_200_OK)

    def import_multi_so(self, request, *args, **kwargs):
        try:
            file = request.FILES.get('file_import', None)
            if not file:
                return Response({'message': f'file_import is required'})
            file_extension = os.path.splitext(file.name)[1].lower()

            if file_extension not in ['.xlsx']:
                return Response({'message': 'File must be .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

            data = self.excel_to_dict(file)
            success, error = self.handle_import_consider(data)
            # Dùng hàm an toàn để convert data thành JSON
            return Response({
                'success': success,
                'errors': error
            })
        except Exception as e:
            raise e
            # return Response({'error': str(e)}, status=500)

    def excel_to_dict(self, file):
        df = pd.read_excel(file, engine='openpyxl')
        column_mapping = {
            'Mã ưu đãi': 'so_id',
            'Mã khách hàng': 'client_id',
            'Tên ưu đãi': 'so_name',
            'Ngày bắt đầu': 'so_start',
            'Ngày kết thúc': 'so_end',
            'Ghi chú': 'note',
            'Tính doanh số': 'count_turnover',
            'Trừ doanh số': 'target',
            'Mã sản phẩm': 'product',
            'Tên sản phẩm': 'product_name',
            'Số lượng': 'quantity_in_box',
            'Số thùng': 'max_order_box',
            'Đơn giá': 'price',
            'Điểm': 'point',
            'Giá trị ưu đãi': 'cashback'
        }
        df.rename(columns=column_mapping, inplace=True)
        df['line_number'] = df.index + 1

        df.replace({np.inf: None, -np.inf: None, np.nan: None}, inplace=True)

        # Chuyển đổi định dạng ngày nếu cần thiết
        df['so_start'] = df['so_start'].apply(convert_date_format)
        df['so_end'] = df['so_end'].apply(convert_date_format)

        # Nhóm dữ liệu theo 'so_id' và chuyển đổi sang định dạng yêu cầu
        grouped_data = df.groupby('so_id')

        result = []
        for so_id, group in grouped_data:
            # Lấy các cột dữ liệu không liên quan đến sản phẩm
            common_data = group.iloc[0][[
                'so_id', 'client_id', 'so_name', 'so_start', 'so_end',
                'note', 'count_turnover', 'target'
            ]].to_dict()

            # Lấy danh sách các sản phẩm
            products = group[
                ['product', 'product_name', 'quantity_in_box', 'max_order_box', 'price',
                 'cashback', 'line_number']].to_dict(orient='records')

            # Ghép dữ liệu chung và danh sách sản phẩm vào một dict
            common_data['so_detail'] = products
            result.append(common_data)

        return result

    def handle_import_consider(self, data):
        try:
            error_data = list()
            update_products_offer = list()
            with transaction.atomic():
                for i, item in enumerate(data):
                    data_lines = [detail['line_number'] for detail in item['so_detail']]

                    try:
                        count_turnover = True if item['count_turnover'] in ['x', 'X'] else False
                        so_name = item['so_name'] if item['so_name'] not in ['', None] else f"Ưu đãi {item['so_id']}"
                        so_obj = SpecialOffer.objects.create(
                            name=so_name,
                            time_start=item['so_start'],
                            time_end=item['so_end'],
                            target=item['target'],
                            count_turnover=count_turnover,
                            note=item['note'],
                            type_list=so_type.consider_user,
                            status=data_status.active
                        )
                    except Exception as e:
                        error = {
                            'line_id': item['so_id'],
                            'message': f'lỗi khi tạo ưu đãi {item["so_name"]}',
                            'error_line': data_lines
                        }
                        error_data.append(error)
                        raise e

                    try:
                        client_id = item['client_id']
                        actions = [perm_actions['view'], perm_actions['create']]
                        list_perm = create_full_perm(SpecialOffer, so_obj.id, actions)

                        # Get user has perm
                        existed_user_allow = list_user_has_perm(list_perm, True)

                        add_perm({'type': 'users', 'data': [client_id], 'existed': existed_user_allow}, list_perm, True)

                    except Exception as e:
                        error = {
                            'line_id': item['so_id'],
                            'message': f'lỗi khi thêm user {client_id} vào ưu đãi {item["so_name"]}',
                            'error_line': data_lines
                        }
                        error_data.append(error)
                        raise e
                    products_offer = list()
                    try:
                        for product in item['so_detail']:
                            product_name = product.pop('product_name', None)
                            product_id = product.pop('product')
                            line_number = product.pop('line_number')

                            product_obj = Product.objects.get(id=product_id)

                            product['special_offer'] = so_obj
                            print(f"Test: {product}")
                            product_price = SpecialOfferProduct(product=product_obj, **product)
                            products_offer.append(product_price)
                    except Product.DoesNotExist:
                        error = {
                            'line_id': item['so_id'],
                            'message': f'không tìm thấy quy cách {product_id}',
                            'error_line': data_lines
                        }
                        error_data.append(error)
                        continue
                    success = {
                        'line_id': item['so_id'],
                        'new_id': so_obj.id,
                    }
                    update_products_offer.append(success)
                    SpecialOfferProduct.objects.bulk_create(products_offer)
                    if i == 5:
                        raise Exception("Break for testing")
            return update_products_offer, error_data
        except Exception as e:
            raise e


def handle_non_serializable(value):
    if isinstance(value, float):
        if math.isinf(value) or math.isnan(value):
            return None  # Hoặc một giá trị phù hợp khác
    return value


class ApiSpecialOfferConsider(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = SpecialOfferSerializer
    queryset = SpecialOffer.objects.filter(type_list=so_type.consider_user)
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=PriceList)]

    def list(self, request, *args, **kwargs):
        user_id = request.query_params.get('user', None)
        user = User.objects.filter(id=user_id)
        if not user.exists():
            return ValidationError({'message': f'user id {user_id} not found'})
        user = user.first()
        queryset = perm_queryset(self, user)
        response = filter_data(self, request, ['id', 'name', 'status'],
                               queryset=queryset, **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiImportProductPL(APIView):
    def post(self, request, *args, **kwargs):
        app_log.info(f"Importing product price list")
        excel_file = self.request.FILES.get('excel_file', None)
        app_log.info(f"Test excel file: {str(excel_file)}")
        # Validate
        if excel_file is None:
            return Response({'message': "not found excel import file"}, status=status.HTTP_400_BAD_REQUEST)
        # Validate file extension
        file_extension = os.path.splitext(excel_file.name)[1].lower()
        if file_extension not in ['.xls', '.xlsx']:
            return Response({'message': "Invalid file format. Only .xls and .xlsx files are supported."},
                            status=status.HTTP_400_BAD_REQUEST)

        price_list_id = self.request.data.get('price_list', None)
        if price_list_id is None:
            return Response({'message': "field price_list is required"}, status=status.HTTP_400_BAD_REQUEST)
        price_list = PriceList.objects.filter(id=price_list_id)
        if not price_list.exists():
            return Response({'message': f"price list with id {price_list_id} not found"},
                            status=status.HTTP_400_BAD_REQUEST)
        price_list = price_list.first()
        try:
            with transaction.atomic():
                excel_data = pandas.read_excel(excel_file, usecols="A:D")

                bulk_create_list = []
                bulk_update_list = []
                product_ids_in_excel = []

                for index, row in excel_data.iterrows():
                    # Get data from excel file
                    product_id = row['maSanPham']
                    product_price = row['donGia']
                    quantity_in_box = row['soLuongTrenThung']
                    point = row['diemTrenThung']
                    # Trying get product id
                    try:
                        product = Product.objects.get(id=product_id)
                    except Product.DoesNotExist:
                        return Response({'message': f"Product with id {product_id} not found"},
                                        status=status.HTTP_400_BAD_REQUEST)

                    product_ids_in_excel.append(product_id)

                    product_price_object = ProductPrice.objects.filter(price_list=price_list, product=product).first()
                    # Update existed product price
                    if product_price_object:
                        product_price_object.price = product_price
                        product_price_object.quantity_in_box = quantity_in_box
                        product_price_object.point = point
                        bulk_update_list.append(product_price_object)
                    # Create objects and append to create list
                    else:
                        product_price_object = ProductPrice(
                            price_list=price_list,
                            product=product,
                            price=product_price,
                            quantity_in_box=quantity_in_box,
                            point=point
                        )
                        bulk_create_list.append(product_price_object)

                # Bulk create new entries
                if bulk_create_list:
                    ProductPrice.objects.bulk_create(bulk_create_list)
                # Bulk update existing entries
                if bulk_update_list:
                    ProductPrice.objects.bulk_update(bulk_update_list, ['price', 'quantity_in_box', 'point'])

                # Remove products not in the current Excel file
                ProductPrice.objects.filter(price_list=price_list).exclude(
                    product__id__in=product_ids_in_excel).delete()

            data = PriceList2Serializer(price_list).data
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            app_log.info(f"Error when updating product price")
            raise e


class ApiImportProductSO(APIView):
    def post(self, request, *args, **kwargs):
        # Get request data
        excel_file = self.request.FILES.get('excel_file', None)
        so_id = self.request.data.get('special_offer', None)
        app_log.info(f"Test excel file: {str(excel_file)}")

        # Validate file
        if excel_file is None:
            return Response({'message': "not found excel import file"}, status=status.HTTP_400_BAD_REQUEST)
        # Validate file extension
        file_extension = os.path.splitext(excel_file.name)[1].lower()
        if file_extension not in ['.xls', '.xlsx']:
            return Response({'message': "invalid file format, only .xls and .xlsx files are supported."},
                            status=status.HTTP_400_BAD_REQUEST)
        # Validate Special Offer
        if so_id is None:
            return Response({'message': "field price_list is required"}, status=status.HTTP_400_BAD_REQUEST)
        special_offer = SpecialOffer.objects.filter(id=so_id).first()
        if not special_offer:
            return Response({'message': f"special offer with id {so_id} not found"},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            so_update = list()
            so_create = list()
            product_in_excel = list()

            with transaction.atomic():
                excel_data = pandas.read_excel(excel_file, usecols="A:F")
                for index, row in excel_data.iterrows():
                    # Get data from excel file
                    product_id = row['maSanPham']
                    product_price = row['donGia']
                    quantity_in_box = row['soLuongTrenThung']
                    point = row['diemTrenThung']
                    cashback = row['hoanTien']
                    max_box = row['thungMuaToiDa']
                    app_log.info(
                        f"Test : {product_id} | {product_price} | {quantity_in_box} | {point} | {cashback} | {max_box}")
                    # Trying get product id
                    try:
                        product = Product.objects.get(id=product_id)
                    except Product.DoesNotExist:
                        return Response({'message': f"Product with id {product_id} not found"},
                                        status=status.HTTP_400_BAD_REQUEST)

                    product_in_excel.append(product_id)
                    so_product = SpecialOfferProduct.objects.filter(special_offer=special_offer,
                                                                    product=product).first()
                    app_log.info(f"test product & so: {product} | {so_product}")
                    if so_product:
                        so_product.price = product_price
                        so_product.quantity_in_box = quantity_in_box
                        so_product.point = point
                        so_product.cashback = cashback
                        so_product.max_order_box = max_box
                        # Add so product to update list
                        app_log.info(f"test product exist: {so_product}")

                        so_update.append(so_product)
                    else:
                        so_product = SpecialOfferProduct(
                            special_offer=special_offer,
                            product=product,
                            price=product_price,
                            quantity_in_box=quantity_in_box,
                            point=point,
                            cashback=cashback,
                            max_order_box=max_box
                        )
                        app_log.info(f"test product create: {so_product}")
                        # Add so product to create list
                        so_create.append(so_product)
                # Bulk create new entries
                app_log.info(f"test so_create: {so_create}")
                SpecialOfferProduct.objects.bulk_create(so_create)
                # Bulk update existing entries
                app_log.info(f"test so_update: {so_update}")
                SpecialOfferProduct.objects.bulk_update(so_update, ['price', 'quantity_in_box', 'point', 'cashback',
                                                                    'max_order_box'])

                # Remove products not in the current Excel file
                SpecialOfferProduct.objects.filter(special_offer=special_offer).exclude(
                    product__id__in=product_in_excel).delete()
            data = SpecialOfferSerializer(special_offer).data
            return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            app_log.error(f"Get error when import data for ")
            raise e


class ApiSOProduct(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = SpecialOfferProductSerializer
    queryset = SpecialOfferProduct.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=SpecialOfferProduct)]

    # def get_queryset(self):
    #     app_log.info(f"Getting query set")
    #     return perm_queryset(self)

    def list(self, request, *args, **kwargs):
        queryset = SpecialOfferProduct.objects.annotate(
            special_offer_created_at=F('special_offer__created_at')
        ).order_by('-special_offer_created_at')
        response = filter_data(self, request,
                               ['special_offer__id', 'special_offer__type_list', 'special_offer__priority',
                                'product__name', 'product__id'],
                               queryset=queryset, order_by_required=False,
                               **kwargs)

        return Response(response, status.HTTP_200_OK)
