import os
import random
import sys
import tempfile
import traceback
from functools import partial
from io import BytesIO
from types import SimpleNamespace

import pandas as pd
from django.conf import settings
from django.db import transaction
from django.db.models import QuerySet, F
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import viewsets, mixins, status
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from weasyprint import HTML

from account.handlers.perms import perm_queryset
from account.handlers.validate_perm import ValidatePermRest
from account.models import User
from app.logs import app_log
from marketing.pick_number.api.serializers import UserJoinEventSerializer, EventNumberSerializer, NumberListSerializer, \
    UserJoinEventNumberSerializer, PrizeEventSerializer, PickNumberLogSerializer, \
    AwardUserSerializer
from marketing.pick_number.models import UserJoinEvent, EventNumber, NumberList, PrizeEvent, AwardNumber, PickNumberLog, \
    NumberSelected
from user_system.employee_profile.models import EmployeeProfile
from utils.model_filter_paginate import filter_data


class ApiEventNumber(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = EventNumberSerializer
    queryset = EventNumber.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=EventNumber)]

    def get_queryset(self):
        user = self.request.user
        return perm_queryset(self, user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        user_id = request.query_params.get('user', None)
        if user_id:
            user = User.objects.filter(id=user_id)
            if not user.exists():
                return Response({'message': f'user id {user_id} not found'})
            user = user.first()
            queryset = perm_queryset(self, user)
        response = filter_data(self, request, ['id', 'name', 'date_start', 'date_close', 'status',
                                               'user_join_event__user__id'], queryset=queryset,
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiUserJoinEvent(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                       mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = UserJoinEventSerializer
    queryset = UserJoinEvent.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=UserJoinEvent)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['id', 'event__id', 'event__name', 'user__id'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiNumberList(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                    mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = NumberListSerializer
    queryset = NumberList.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=NumberList)]

    def list(self, request, *args, **kwargs):
        active = request.query_params.get('active', 0)
        if active == '1':
            queryset = self.queryset.filter(repeat_count__gt=0)
        elif active == '0':
            queryset = self.queryset.filter(repeat_count=0)
        else:
            queryset = self.queryset
        response = filter_data(self, request, ['id', 'event__id', 'event__name'], queryset=queryset,
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiPickNumber(viewsets.GenericViewSet, mixins.RetrieveModelMixin, mixins.UpdateModelMixin):
    serializer_class = UserJoinEventNumberSerializer
    queryset = UserJoinEvent.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=UserJoinEvent)]


class ApiAwardNumber(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = AwardUserSerializer
    queryset = AwardNumber.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=AwardNumber)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request,
                               ['id', 'prize__prize_name', 'prize__event__name', 'prize__prize_id', 'prize__event__id',
                                'number'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiPrizeEvent(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                    mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = PrizeEventSerializer
    queryset = PrizeEvent.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=PrizeEvent)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['event__name', 'event__id', 'prize_name', 'note', 'id'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)

    def import_prize_number(self, request, *args, **kwargs):
        file = request.FILES.get('file_import', None)
        if not file:
            return Response({'message': f'file_import is required'})
        file_extension = os.path.splitext(file.name)[1].lower()

        if file_extension not in ['.xlsx']:
            return Response({'message': 'File must be .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            return Response({'message': f'Error reading the Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        column_mapping = {
            'idGiai': 'prize_id',
            'soTrungGiai': 'number',
            'luotQuay': 'turn_roll'
        }
        missing_columns = [col for col in column_mapping.keys() if col not in df.columns]
        if missing_columns:
            return Response({'message': f'lỗi file không chứa cột: {", ".join(missing_columns)}'},
                            status=status.HTTP_400_BAD_REQUEST)

        df.rename(columns=column_mapping, inplace=True)
        df.replace({pd.NA: None, pd.NaT: None}, inplace=True)
        df['line_number'] = df.index + 1
        prize_number_data = df.to_dict(orient='records')

        success, errors = self.handle_prize_number(prize_number_data)

        return Response({'message': 'ok', 'success': success, 'errors': errors})

    def handle_prize_number(self, prize_number_data):
        success = list()
        errors = list()
        added = list()
        try:
            with transaction.atomic():
                ids = {f"{prize_num['prize_id']}" for prize_num in prize_number_data}
                available_prizes = PrizeEvent.objects.filter(id__in=ids)
                prize_dict = {}
                for prize in available_prizes:
                    prize_data = dict()
                    prize_data['award_numbers_count'] = prize.award_number.all().count()
                    prize_data['reward_number'] = prize.reward_number
                    prize_dict[prize.id] = prize_data
                available_ids = available_prizes.values_list('id', flat=True).distinct()
                for prize_num in prize_number_data:
                    print(f"Test data: {prize_num}")
                    line_number = prize_num.pop('line_number')
                    if prize_num['prize_id'] not in available_ids:
                        error_data = {
                            "line": line_number,
                            "message": f"{prize_num['prize_id']} id giải thưởng không tồn tại"
                        }
                        errors.append(error_data)
                        continue
                    prize = prize_dict.get(prize_num['prize_id'])
                    if prize['award_numbers_count'] >= prize['reward_number']:
                        error_data = {
                            "line": line_number,
                            "message": f"số lượng số trứng thưởng vượt quá giới hạn {prize['reward_number']}"
                        }
                        errors.append(error_data)
                        continue
                    new_prize_number = AwardNumber.objects.create(**prize_num)
                    success.append({
                        "line": line_number,
                        "message": new_prize_number.id
                    })
                    added.append(new_prize_number)

                return success, errors
        except Exception as e:
            if settings.DEBUG:
                raise e
            else:
                error_trace = traceback.format_exc()
                app_log.error(f"Error when handle import prize number: "
                              f"{str(e)}\nDetails: {error_trace}")


class ApiPickNumberLog(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = PickNumberLogSerializer
    queryset = PickNumberLog.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=PickNumberLog)]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        event = request.query_params.get('event', None)
        numbers = request.query_params.get('numbers', None)

        if event:
            queryset = queryset.filter(event__id=event)
        if numbers:
            numbers_list = numbers.split(',')
            queryset = queryset.filter(number__in=numbers_list)

        response = filter_data(self, request, ['event__name', 'event__id', 'user__id',
                                               'user__username',
                                               'user__clientprofile__register_name', 'action', 'phone__phone_number'],
                               queryset=queryset,
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiExportEventNumber(APIView):
    def get(self, request, *args, **kwargs):
        event_id = self.kwargs.get('pk')
        event = EventNumber.objects.filter(id=event_id).select_related('table_point').first()
        if event is None:
            return Response({'message': 'Event ID is required'}, status=400)

        # Tạo workbook và các sheet
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Báo cáo sự kiện"
        ws2 = wb.create_sheet(title="Danh sách số")

        # Xử lý và ghi dữ liệu vào sheet 1
        self.process_event_data(event, ws1)

        # Xử lý và ghi dữ liệu vào sheet 2
        number_lists = event.number_list.all().order_by('number')
        data = [
            {'Tem số': nl.number, 'Tem dư': nl.repeat_count} for nl in number_lists
        ]
        df = pd.DataFrame(data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws2.append(r)

        # Lưu workbook vào BytesIO stream và trả về response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="report_{event_id}.xlsx"'

        return response

    def process_event_data(self, event, ws):
        # Định nghĩa font và border
        header_font = Font(size=13, bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Tạo header
        headers = ["Mã KH", "Số đã chọn"]
        ws.append(headers)

        # Định dạng header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.column_dimensions['B'].width = 13
        # Thêm dữ liệu
        users_join_event = UserJoinEvent.objects.filter(event=event).select_related('user').prefetch_related(
            'number_selected__number')

        for user_join_event in users_join_event:
            user = user_join_event.user
            selected_numbers = user_join_event.number_selected.all().order_by('created_at')
            for number_selected in selected_numbers:
                ws.append([user.id, number_selected.number.number])


class ApiExportEventNumberUser(APIView):
    def get(self, request, *args, **kwargs):
        event_id = self.kwargs.get('pk')
        event = EventNumber.objects.filter(id=event_id).select_related('table_point').first()
        if event is None:
            return Response({'message': 'event id is required'}, status=400)

        # Tạo workbook và worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo sự kiện"

        # Định nghĩa font và border
        title_font = Font(size=12, bold=True, underline="single")
        header_font = Font(size=14, bold=True)
        data_font = Font(size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Tạo tiêu đề cho báo cáo
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Báo cáo sự kiện {event.name}"
        title_cell.font = title_font
        # title_cell.alignment = center_alignment

        ws.append([])  # Thêm dòng trống

        # Tạo header
        headers = ["Mã KH", "Tên KH",
                   "Mã NVTT", "Tên NVTT",
                   "Tổng tem đạt", "Tem chưa chọn", "Số đã chọn"]
        ws.append(headers)

        # Định dạng header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        ws.column_dimensions['A'].width = 12.73  # Đặt độ rộng cột "Mã KH"
        ws.column_dimensions['B'].width = 20

        ws.column_dimensions['C'].width = 12.73
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12.73
        ws.column_dimensions['F'].width = 12.73
        ws.column_dimensions['G'].width = 32

        # Lấy danh sách user_join_event với prefetch_related cho NumberSelected
        users_join_event = UserJoinEvent.objects.filter(event=event).select_related('user').prefetch_related(
            'number_selected__number')
        nvtt_ids = {user_join_event.user.clientprofile.nvtt_id for user_join_event in users_join_event}

        nvtt_profiles = EmployeeProfile.objects.filter(employee_id__in=nvtt_ids).values_list('employee_id_id', 'register_name').distinct()
        nvtt_profiles = dict(nvtt_profiles)
        row_num = 4  # Bắt đầu từ hàng thứ 4 do tiêu đề và header đã chiếm 3 hàng
        for user_join_event in users_join_event:
            user = user_join_event.user
            # print(f"Test: {nvtt_profiles[user.clientprofile.nvtt_id]}")
            selected_numbers: QuerySet = user_join_event.number_selected.all().order_by('created_at')
            turn_pick = user_join_event.turn_pick or 0
            turn_not_pick = turn_pick - selected_numbers.count()

            try:
                register_name = user.clientprofile.register_name
            except AttributeError:
                register_name = ''

            picked_numbers = selected_numbers.values_list('number__number', flat=True).distinct()
            split_numbers = ",".join(map(str, list(picked_numbers)))
            try:
                nvtt_name = nvtt_profiles[user.clientprofile.nvtt_id]
            except Exception:
                nvtt_name = ''
            export_data = [user.id, register_name, user.clientprofile.nvtt_id,
                           nvtt_name, turn_pick, turn_not_pick, split_numbers]
            self._write_to_sheet(ws, row_num, export_data, data_font, center_alignment, thin_border)
            row_num += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Tạo HttpResponse để trả về file Excel
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Bao_cao_su_kien_{event_id}.xlsx'

        return response

    def _write_to_sheet(self, worksheet, row_num, data, font, alignment, border):
        """Helper function to write a row of data to the worksheet with formatting."""
        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = font
            if col_num == 2:  # Căn giữa cột "Số đã chọn"
                cell.alignment = alignment
            cell.border = border


class ApiUserAward(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = AwardUserSerializer
    queryset = AwardNumber.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]

    permission_classes = [partial(ValidatePermRest, model=PickNumberLog)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['number'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)

    def export_award(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        evn = EventNumber.objects.filter(id=pk)
        if not evn.exists():
            return Response({'message': f"không tìm thấy sự kiện quay số {pk}"}, 404)



class ApiUserNumberPdf2(APIView):
    def get(self, request, pk):
        try:
            # Lấy thông tin người dùng và sự kiện dựa trên primary key
            user_event = UserJoinEvent.objects.select_related('user__clientprofile', 'event').get(pk=pk)
            event = user_event.event

            # Lấy danh sách các số đã chọn
            numbers_selected = NumberSelected.objects.filter(user_event=user_event).order_by(
                'number__number').values_list('number__number', flat=True)
            if user_event.turn_pick is None:
                user_event.turn_pick = 0
            if user_event.turn_per_point is None:
                user_event.turn_per_point = 0
            point_redundant = user_event.total_point - (user_event.turn_pick * user_event.turn_per_point)
            if point_redundant < 0:
                point_redundant = 0
            turn_not_pick = user_event.turn_pick - user_event.turn_selected
            print(f"Test turn not pick: {turn_not_pick}")
            # Tạo context để render template
            context = {
                'data': {
                    'event_name': event.name,
                    'username': user_event.user.clientprofile.register_name,
                    'usercode': user_event.user.id,
                    'turn_roll': user_event.turn_pick,
                    'turn_chosen': user_event.turn_selected,
                    'turn_not_pick': turn_not_pick,
                    'total_point': user_event.total_point,
                    'user_point': point_redundant,  # Giả sử 'total_point' là 'user_point'
                    'number_rolled_str': ', '.join(map(str, numbers_selected)),
                },
                'number_rolled': numbers_selected,
            }

            # Render template HTML thành chuỗi
            html_string = render_to_string('pdfs/user_numbers.html', context)

            # Tạo một tệp PDF tạm thời
            with tempfile.NamedTemporaryFile(delete=True) as output:
                HTML(string=html_string).write_pdf(output.name)
                # Đọc PDF
                output.seek(0)
                pdf = output.read()

            # Thiết lập HttpResponse
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{user_event.user.username}_number.pdf"'
            return response
        except UserJoinEvent.DoesNotExist:
            return Response({'error': 'UserJoinEvent not found'}, status=status.HTTP_404_NOT_FOUND)


class ApiUserNumberPdf(APIView):
    def get(self, request, pk):
        try:
            # Lấy thông tin người dùng và sự kiện dựa trên primary key
            user_event = UserJoinEvent.objects.select_related('user__clientprofile', 'event').get(pk=pk)
            event = user_event.event

            # Lấy danh sách các số đã chọn
            numbers_selected = NumberSelected.objects.filter(user_event=user_event).order_by(
                'number__number').values_list('number__number', flat=True)
            if user_event.turn_pick is None:
                user_event.turn_pick = 0
            if user_event.turn_per_point is None:
                user_event.turn_per_point = 0

            point_redundant = user_event.total_point - (user_event.turn_pick * user_event.turn_per_point)
            if point_redundant < 0:
                point_redundant = 0
            turn_not_pick = user_event.turn_pick - user_event.turn_selected
            print(f"Test turn not pick: {turn_not_pick}")
            # Tạo context để render template
            context = {
                'request': request,
                'data': {
                    'event_name': event.name,
                    'username': user_event.user.clientprofile.register_name,
                    'usercode': user_event.user.id,
                    'turn_roll': user_event.turn_pick,
                    'turn_chosen': user_event.turn_selected,
                    'turn_not_pick': turn_not_pick,
                    'total_point': user_event.total_point,
                    'user_point': point_redundant,  # Giả sử 'total_point' là 'user_point'
                    'number_rolled_str': ', '.join(map(str, numbers_selected)),
                },
                'number_rolled': numbers_selected,
            }

            # Render template HTML thành chuỗi
            html_string = render_to_string('pdfs/user_numbers.html', context)

            # Sử dụng BytesIO để tạo PDF
            pdf_file = BytesIO()
            HTML(string=html_string).write_pdf(pdf_file)
            pdf_file.seek(0)  # Quay về đầu file để chuẩn bị gửi dữ liệu

            # Thiết lập HttpResponse
            response = HttpResponse(pdf_file.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{user_event.user.username}_number.pdf"'
            return response
        except UserJoinEvent.DoesNotExist:
            return Response({'error': 'UserJoinEvent not found'}, status=status.HTTP_404_NOT_FOUND)


class RandomNumber(APIView):
    def post(self, request, pk):
        try:
            user_event: UserJoinEvent = UserJoinEvent.objects.get(pk=pk)
            event: EventNumber = user_event.event
            try:
                random_times = request.data.get('random_times', 1)
                random_times = int(random_times)
            except TypeError:
                return Response({'message': 'random_times phải là một số integer'}, 400)
            rand_nums, rand_nums_obj = get_random_numbers(user_event, random_times)
            handle_selecting(user_event, rand_nums_obj)
            selected_numbers: QuerySet[NumberSelected] = user_event.number_selected.filter().values_list(
                'number__number', flat=True)
            response = {
                'user': user_event.user.id,
                'event': event.id,
                'user_event': user_event.id,
                'random_number': rand_nums,
                'selected_number': list(selected_numbers)
            }

            return Response(response, 200)
        except Exception as e:
            raise e


class ApiAutoPick(APIView):
    def post(self, request, pk):
        try:
            event: EventNumber = EventNumber.objects.filter(pk=pk).first()
            add_more_number = request.data.get('add_more_number', False)
            if not event:
                return Response({'message': f'not found {pk}'}, 404)
            users_event = event.user_join_event.annotate(
                turns_left=F('turn_pick') - F('turn_selected')
            ).filter(turns_left__gt=0)
            for user_event in users_event:
                random_times = user_event.turn_pick - user_event.turn_selected
                rand_nums, rand_nums_obj = get_random_numbers(user_event, random_times)
                first_loop = True
                while add_more_number and len(rand_nums) < random_times and len(rand_nums_obj) < random_times:
                    if first_loop:
                        event.range_number += round((random_times - len(rand_nums_obj)) * 0.9)
                    else:
                        event.range_number += 1
                    event.save()
                    rand_nums, rand_nums_obj = get_random_numbers(user_event, random_times)
                    if len(rand_nums) >= random_times and len(rand_nums_obj) >= random_times:
                        break
                handle_selecting(user_event, rand_nums_obj)

            return Response({'message': 'success'})
        except Exception as e:
            raise e


def handle_selecting(user_event, rand_nums_obj):
    try:
        with transaction.atomic():
            numbers_selected = list()
            saved_logs = list()
            for rand_num_obj in rand_nums_obj:

                while rand_num_obj.repeat_count <= 0:
                    rand_nums, rand_num_obj = get_random_numbers(user_event, 1)
                    if rand_num_obj.repeat_count > 0:
                        break
                new_added = NumberSelected(user_event=user_event, number=rand_num_obj)
                add_log = PickNumberLog(event=user_event.event, user=user_event.user, number=rand_num_obj.number,
                                        action='random')
                saved_logs.append(add_log)
                numbers_selected.append(new_added)
                rand_num_obj.repeat_count -= 1
                rand_num_obj.save()
            NumberSelected.objects.bulk_create(numbers_selected)
            PickNumberLog.objects.bulk_create(saved_logs)
            selected_numbers = user_event.number_selected.filter().values_list(
                'number__number', flat=True).count()
            print(f"Test: {user_event}")
            # user_event = user_event.first()
            user_event.turn_selected = selected_numbers
            user_event.save()
    except Exception as e:
        raise e


def get_random_numbers(user_event: UserJoinEvent, random_times: int = 1):
    event = user_event.event
    selected_number = NumberSelected.objects.filter(user_event=user_event).values_list('number__number', flat=True)
    available_number = event.number_list.filter(repeat_count__gt=0).exclude(number__in=selected_number).values_list(
        'number', flat=True).distinct()
    available_number = list(set(available_number))
    if len(available_number) == 0:
        return [], []
    if random_times > len(available_number):
        random_times = len(available_number)
    random_elements = random.sample(available_number, random_times)

    valid_numbers = []
    for num in random_elements:
        if event.number_list.filter(number=num, repeat_count__gt=0).exists():
            valid_numbers.append(num)

    while len(valid_numbers) < random_times:
        remaining_count = random_times - len(valid_numbers)
        available_number = event.number_list.filter(repeat_count__gt=0).exclude(number__int=selected_number).exclude(
            number__in=valid_numbers).values_list(
            'number', flat=True).distinct()
        available_number = list(set(available_number))
        if not available_number:
            break
        if remaining_count > len(available_number):
            remaining_count = len(available_number)
        number_list = random.sample(available_number, remaining_count)
        for num in number_list:
            if event.number_list.filter(number=num, repeat_count__gt=0).exists():
                valid_numbers.append(num)
    number_objs = event.number_list.filter(number__in=valid_numbers)
    return valid_numbers, number_objs
