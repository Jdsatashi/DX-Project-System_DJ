from functools import partial

import openpyxl
import pandas as pd
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, mixins, status
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

from account.handlers.validate_perm import ValidatePermRest
from account.models import User
from app.logs import app_log
from marketing.order.models import OrderDetail
from marketing.sale_statistic.api.serializers import SaleStatisticSerializer, SaleMonthTargetSerializer, \
    UserSaleStatisticSerializer, UserUsedStatsSerializer
from marketing.sale_statistic.models import SaleStatistic, SaleTarget, UserSaleStatistic, UsedTurnover
from system_func.models import PeriodSeason
from utils.helpers import local_time
from utils.model_filter_paginate import filter_data


class ApiSaleStatistic(viewsets.GenericViewSet, mixins.ListModelMixin,
                       mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = SaleStatisticSerializer
    queryset = SaleStatistic.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=SaleStatistic)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['user__id'],
                               **kwargs)
        return Response(response, status=status.HTTP_200_OK)


class CurrentMonthSaleStatisticView(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = SaleStatisticSerializer
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=SaleTarget)]

    def get_queryset(self):
        today = timezone.now().date()
        first_day_of_month = today.replace(day=1)
        return SaleStatistic.objects.filter(month=first_day_of_month)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        response = filter_data(self, request, ['user__id'], queryset=queryset,
                               **kwargs)
        return Response(response, status=status.HTTP_200_OK)


class UserMonthSaleStatisticView(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = SaleStatisticSerializer
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=SaleTarget)]

    def get_queryset(self):
        user_id = self.kwargs.get('user_id')
        user = User.objects.get(id=user_id)

        return SaleStatistic.objects.filter(user=user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        response = filter_data(self, request, ['month'], queryset=queryset,
                               **kwargs)
        return Response(response, status=status.HTTP_200_OK)


class ApiSaleMonthTarget(viewsets.GenericViewSet, mixins.ListModelMixin,
                         mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = SaleMonthTargetSerializer
    queryset = SaleTarget.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=SaleTarget)]

    def list(self, request, *args, **kwargs):
        today = local_time().date()
        sale_target, _ = SaleTarget.objects.get_or_create(month=today)
        app_log.info(f"{sale_target}")
        response = filter_data(self, request, ['month'],
                               **kwargs)
        return Response(response, status=status.HTTP_200_OK)


class ApiMainSaleStatistic(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin):
    serializer_class = UserSaleStatisticSerializer
    queryset = UserSaleStatistic.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=UserSaleStatistic)]

    def list(self, request, *args, **kwargs):

        response = filter_data(self, request, ['user__id', 'user__username'],
                               **kwargs)
        return Response(response, status=status.HTTP_200_OK)

    def import_file(self, request, *args, **kwargs):
        file = request.data.get('import_file', None)
        if file is None:
            return Response({'message': 'import_file is required'})
        try:
            with transaction.atomic():
                df = pd.read_excel(file, engine='openpyxl')

                if 'maKH' not in df.columns or 'thay_doi_doanh_so' not in df.columns:
                    return Response({'message': 'File phải chứa các cột "maKH" và "thay_doi_doanh_so"'})

                data = df[['maKH', 'thay_doi_doanh_so', 'ghi_chu']]
                update_turnover = list()
                create_record = list()
                for index, row in data.iterrows():
                    user_id = row['maKH']
                    fix_turnover = row['thay_doi_doanh_so']
                    user_stats = UserSaleStatistic.objects.filter(user=user_id).first()
                    if user_stats is None:
                        user = User.objects.get(id=user_id)
                        user_stats = UserSaleStatistic.objects.create(user=user)
                    user_stats.turnover += fix_turnover
                    update_turnover.append(user_stats)
                    note = row['ghi_chu']
                    if note in ['nan', None, '']:
                        note = 'import excel'
                    record = UsedTurnover(user_sale_stats=user_stats, turnover=fix_turnover, purpose='admin_fix', note=note)
                    create_record.append(record)
                UserSaleStatistic.objects.bulk_update(update_turnover, ['turnover'])
                UsedTurnover.objects.bulk_create(create_record)
                return Response({'message': 'Import file successfully'})

        except Exception as e:
            raise e

    def export_file(self, request, *args, **kwargs):
        data = UserSaleStatistic.objects.all().order_by('user__id').values('user__id', 'turnover')

        # Tạo Workbook mới
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "User Sale Statistics"

        # Ghi tiêu đề cột
        columns = ['MaKH', 'Doanh Số', 'Thùng ưu đãi đã dùng']
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            sheet[f'{column_letter}1'] = column_title

        current_season: PeriodSeason = PeriodSeason.get_period_by_date('turnover')

        for row_num, instance in enumerate(data, 2):
            user_id = instance['user__id']
            turnover = instance['turnover']

            # Lấy đối tượng User
            user = User.objects.get(id=user_id)

            # Lấy dữ liệu used_box cho user
            user_so = user.order_set.filter(
                is_so=True,
                date_get__gte=current_season.from_date,
                date_get__lte=current_season.to_date
            )

            used_box = OrderDetail.objects.filter(order_id__in=user_so).aggregate(total_box=Sum('order_box'))

            sheet[f'A{row_num}'] = user.id  # Cột MaKH
            sheet[f'B{row_num}'] = turnover  # Cột Doanh Số
            sheet[f'C{row_num}'] = used_box['total_box'] or 0  # Cột Used Box

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_sale_statistics.xlsx'

        workbook.save(response)

        return response


class ApiUserUsedStatistic(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = UserUsedStatsSerializer
    queryset = UsedTurnover.objects.all()
    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=UsedTurnover)]

    def list(self, request, *args, **kwargs):

        response = filter_data(self, request, ['user_sale_stats__user__id', 'user_sale_stats__user__username'],
                               **kwargs)
        return Response(response, status=status.HTTP_200_OK)
