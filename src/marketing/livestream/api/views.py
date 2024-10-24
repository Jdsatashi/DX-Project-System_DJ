from functools import partial

import openpyxl
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from rest_framework import mixins, viewsets, status
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication

from account.handlers.perms import perm_queryset, export_users_has_perm
from account.handlers.validate_perm import ValidatePermRest
from account.models import PhoneNumber, User
from app.logs import app_log
from marketing.livestream.api.serializers import LiveStreamSerializer, LiveStreamCommentSerializer, \
    LiveStatistic, LiveTracking, LiveStreamDetailCommentSerializer, PeekViewSerializer, LiveOfferRegisterSerializer
from marketing.livestream.models import LiveStream, LiveStreamComment, LiveStreamTracking, LiveStreamStatistic, \
    LiveStreamPeekView, LiveStreamOfferRegister
from utils.model_filter_paginate import filter_data


class ApiLiveStream(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                    mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = LiveStreamSerializer
    queryset = LiveStream.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication, SessionAuthentication]
    permission_classes = [partial(ValidatePermRest, model=LiveStream)]

    def get_queryset(self):
        user = self.request.user
        return perm_queryset(self, user)

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['title', 'date_released', 'live_url'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)

    def export_users(self, request, pk, *args, **kwargs):
        livestream = LiveStream.objects.filter(id=pk).first()
        if not livestream:
            return Response({'message': f'not found livestream id {pk}'})

        workbook = export_users_has_perm(livestream, pk)

        # Chuẩn bị response trả về
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="UserLivestream_{pk}.xlsx"'

        # Lưu workbook vào response
        workbook.save(response)

        return response


class ApiLiveStreamComment(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin):
    serializer_class = LiveStreamCommentSerializer
    queryset = LiveStreamComment.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=LiveStreamComment)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['live_stream__title', 'viewers__id', 'comments'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiLiveStreamDetailComment(viewsets.GenericViewSet, mixins.ListModelMixin):
    serializer_class = LiveStreamDetailCommentSerializer
    queryset = LiveStreamComment.objects.all()

    def list(self, request, *args, **kwargs):
        livestream_id = self.kwargs.get('livestream_id')
        if not livestream_id:
            return Response({'message': 'livestream_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset().filter(live_stream_id=livestream_id)
        response = filter_data(self, request, ['comment', 'user__username', 'user__id', 'phone__phone_number'],
                               queryset=queryset)
        return Response(response, status=status.HTTP_200_OK)


class ApiLiveStatistic(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                       mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = LiveStatistic
    queryset = LiveStreamStatistic.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=LiveStreamStatistic)]

    def list(self, request, *args, **kwargs):
        live_id = request.query_params.get('live_id', None)
        queryset = self.get_queryset()
        if live_id:
            queryset = queryset.filter(live_stream_id=live_id)
        response = filter_data(self, request, ['live_stream__title', 'id'], queryset=queryset,
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiLiveTracking(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                      mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = LiveTracking
    queryset = LiveStreamTracking.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=LiveStreamTracking)]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        live_id = request.query_params.get('live_id', None)
        if live_id:
            queryset = queryset.filter(live_stream=live_id)
        response = filter_data(self, request, ['live_stream__title', 'phone__phone_number'],
                               queryset=queryset, **kwargs)
        return Response(response, status.HTTP_200_OK)


class ApiPeekView(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                  mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = PeekViewSerializer
    queryset = LiveStreamPeekView.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=LiveStreamComment)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['live_stream__id'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class JoinPeekView(APIView):
    def post(self, request, *args, **kwargs):
        live_stream_id = request.data.get('live_stream_id')
        livestream = LiveStream.objects.filter(id=live_stream_id).first()
        app_log.info(livestream)
        peek_view = LiveStreamPeekView.objects.filter(live_stream=livestream).first()
        app_log.info(peek_view)
        peek_view.in_livestream += 1
        peek_view.save()
        return Response(PeekViewSerializer(peek_view).data, status=status.HTTP_200_OK)


class LeavePeekView(APIView):
    def post(self, request, *args, **kwargs):
        live_stream_id = request.data.get('live_stream_id')
        livestream = LiveStream.objects.filter(id=live_stream_id).first()
        peek_view = LiveStreamPeekView.objects.filter(live_stream=livestream).first()
        app_log.info(peek_view)
        peek_view.out_livestream += 1
        peek_view.save()
        return Response(PeekViewSerializer(peek_view).data, status=status.HTTP_200_OK)


class ApiLiveOfferRegister(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin,
                           mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin):
    serializer_class = LiveOfferRegisterSerializer
    queryset = LiveStreamOfferRegister.objects.all()

    authentication_classes = [JWTAuthentication, BasicAuthentication]

    # permission_classes = [partial(ValidatePermRest, model=LiveStreamComment)]

    def list(self, request, *args, **kwargs):
        response = filter_data(self, request, ['live_stream__title', 'live_stream__id', 'phone__phone_number'],
                               **kwargs)
        return Response(response, status.HTTP_200_OK)


class CheckLiveStreamRegistrationView(APIView):
    def post(self, request, *args, **kwargs):
        live_stream_id = request.data.get('live_stream_id')
        phone_number = request.data.get('phone_number')

        if not live_stream_id or not phone_number:
            return Response({'error': 'live_stream_id and phone_number are required'},
                            status=status.HTTP_400_BAD_REQUEST)

        live_stream = get_object_or_404(LiveStream, id=live_stream_id)
        phone = get_object_or_404(PhoneNumber, phone_number=phone_number)
        user = phone.user

        # Kiểm tra tất cả các số điện thoại của người dùng
        phone_numbers = PhoneNumber.objects.filter(user=user)

        registered = LiveStreamOfferRegister.objects.filter(live_stream=live_stream, phone__in=phone_numbers,
                                                            register=True).exists()

        if registered:
            return Response({'register': True}, status=status.HTTP_200_OK)
        else:
            return Response({'register': False},
                            status=status.HTTP_200_OK)


class ExportLiveReport(APIView):
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        livestream = LiveStream.objects.filter(id=pk).first()
        if not livestream:
            return Response({'message': f'not found live stream {pk}'})

        # Tạo workbook và các sheet
        workbook = openpyxl.Workbook()
        general_sheet = workbook.active
        general_sheet.title = "Tổng quát"
        viewer_sheet = workbook.create_sheet("Người xem")
        order_sheet = workbook.create_sheet("Khách đặt hàng")
        comment_sheet = workbook.create_sheet("Bình luận")

        # Định dạng tiêu đề
        title_font = Font(bold=True, size=14)
        center_aligned_text = Alignment(horizontal='center')

        # Tên cột và kích thước cho sheet "Tổng quát"
        columns_general = ['Tên sự kiện', 'Thời gian diễn ra', 'Link video', 'Tổng lượt xem', 'Tổng số bình luận',
                           'Tổng số đặt hàng']
        column_widths_general = [30, 18, 12, 16, 18.5, 18]

        general_sheet.append(columns_general)
        for i, column in enumerate(columns_general, 1):
            cell = general_sheet.cell(row=1, column=i)
            cell.font = title_font
            cell.alignment = center_aligned_text
            general_sheet.column_dimensions[get_column_letter(i)].width = column_widths_general[i - 1]

        # Dữ liệu cho sheet "Tổng quát"
        statistics = LiveStreamStatistic.objects.filter(live_stream=livestream).first()
        general_sheet.append([
            livestream.title,
            f"{livestream.date_released.strftime('%Y-%m-%d')} {livestream.time_start.strftime('%H:%M')}",
            livestream.live_url,
            statistics.viewers if statistics else 0,
            statistics.comments if statistics else 0,
            statistics.order_times if statistics else 0
        ])

        # Cột và kích thước cho sheet "Khách đặt hàng"
        columns_order = ['SĐT', 'Mã KH', 'Tên KH']
        column_widths_order = [12, 10, 26]

        order_sheet.append(columns_order)
        for i, column in enumerate(columns_order, 1):
            cell = order_sheet.cell(row=1, column=i)
            cell.font = title_font
            cell.alignment = center_aligned_text
            order_sheet.column_dimensions[get_column_letter(i)].width = column_widths_order[i - 1]

        # Dữ liệu cho sheet "Khách đặt hàng"
        for register in LiveStreamOfferRegister.objects.filter(live_stream=livestream).select_related('phone',
                                                                                                      'phone__user'):
            user = register.phone.user
            user_name = user.clientprofile.register_name if hasattr(user, 'clientprofile') else ""
            order_sheet.append([
                register.phone.phone_number,
                user.id,
                user_name
            ])

        # Cột và kích thước cho sheet "Người xem"
        columns_viewer = ['Người xem', 'Mã user', 'Tên user', 'Tổng thời gian theo dõi']
        column_widths_viewer = [14, 18, 24, 25]

        viewer_sheet.append(columns_viewer)
        for i, column in enumerate(columns_viewer, 1):
            cell = viewer_sheet.cell(row=1, column=i)
            cell.font = title_font
            cell.alignment = center_aligned_text
            viewer_sheet.column_dimensions[get_column_letter(i)].width = column_widths_viewer[i - 1]

        # Dữ liệu cho sheet "Người xem"
        phone_times = LiveStreamTracking.objects.filter(live_stream=livestream).values(
            'phone__phone_number', 'phone__user__id').annotate(total_time=Sum('time_watch')).order_by('phone__phone_number')
        for phone_time in phone_times:
            user = User.objects.filter(id=phone_time['phone__user__id']).first()
            if user:
                if hasattr(user, 'clientprofile') and user.clientprofile:
                    user_name = user.clientprofile.register_name
                elif hasattr(user, 'employeeprofile') and user.employeeprofile:
                    user_name = user.employeeprofile.register_name
                else:
                    user_name = ""
            else:
                user_name = ""
            viewer_sheet.append(
                [phone_time['phone__phone_number'],
                 phone_time['phone__user__id'],
                 user_name,
                 str(phone_time['total_time'])
                 ])

        columns_comment = ['SĐT', 'Người dùng', 'Bình luận', 'Thời gian bình luận']
        column_widths_comment = [15, 15, 50, 20]  # Kích thước cột tùy chỉnh

        comment_sheet.append(columns_comment)
        for i, column in enumerate(columns_comment, 1):
            cell = comment_sheet.cell(row=1, column=i)
            cell.font = title_font
            cell.alignment = center_aligned_text
            comment_sheet.column_dimensions[get_column_letter(i)].width = column_widths_comment[i - 1]

        comments = LiveStreamComment.objects.filter(live_stream=livestream).select_related('user', 'phone').order_by('-created_at')
        for comment in comments:
            phone_number = comment.phone.phone_number if comment.phone else ""
            user_id = comment.user.id if comment.user else ""
            comment_text = comment.comment
            comment_time = comment.created_at.strftime('%Y-%m-%d %H:%M:%S')
            comment_sheet.append([phone_number, user_id, comment_text, comment_time])

        apply_data_font(general_sheet, 2, general_sheet.max_row, general_sheet.max_column)
        apply_data_font(order_sheet, 2, order_sheet.max_row, order_sheet.max_column)
        apply_data_font(viewer_sheet, 2, viewer_sheet.max_row, viewer_sheet.max_column)
        apply_data_font(comment_sheet, 2, comment_sheet.max_row, comment_sheet.max_column)

        # Lưu và trả về file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="livestream_{pk}_report.xlsx"'
        workbook.save(response)

        return response


def apply_data_font(sheet, row_start, row_end, column_end):
    data_font = Font(size=12)
    for row in sheet.iter_rows(min_row=row_start, max_row=row_end, max_col=column_end):
        for cell in row:
            cell.font = data_font
