import openpyxl

from openpyxl.utils import get_column_letter
from django.contrib.contenttypes.models import ContentType
from django.db import models
from django.db.models import Q, Model
from rest_framework import status
from rest_framework.response import Response

from account.models import Perm, User, GroupPerm, GroupPermPerms
from app.logs import app_log
from utils.constants import perm_actions


def get_action(view, method):
    # Get action of function
    return {
        'GET': 'view',
        'POST': 'create',
        'PUT': 'update',
        'PATCH': 'update',
        'DELETE': 'destroy'
    }.get(method, 'view')


def get_perm_name(model):
    content_type = ContentType.objects.get_for_model(model)
    return f"{content_type.app_label}_{content_type.model}"


def get_required_permission(model, view, request):
    action = get_action(view, request.method)
    perm_name = get_perm_name(model)
    return f"{action}_{perm_name}"


class DataFKModel:
    def __init__(self, model):
        self.model = model

    def get_fk_fields(self):
        """
        Get all ForeignKey fields of model
        """
        foreign_keys = []
        for field in self.model._meta.get_fields():
            if isinstance(field, (models.ForeignKey, models.OneToOneField)):
                foreign_keys.append(field.name)
        return foreign_keys

    def get_fk_models(self):
        """
        Get all ForeignKey models
        """
        foreign_keys = []
        for field in self.model._meta.get_fields():
            if isinstance(field, (models.ForeignKey, models.OneToOneField)):
                foreign_keys.append(field)
        return foreign_keys

    def get_fk_fields_models(self):
        """
        Get all ForeignKey fields of model
        """
        foreign_keys = []
        for field in self.model._meta.get_fields():
            if isinstance(field, (models.ForeignKey, models.OneToOneField)):
                data = {'field': field.name, 'model': field.related_model}
                foreign_keys.append(data)
        return foreign_keys


def perm_queryset(self, user):
    model_class = self.serializer_class.Meta.model
    pk = self.kwargs.get('pk')
    if user.is_superuser:
        return model_class.objects.all()
    if not user.is_authenticated:
        return Response({'message': 'User is not authenticate'}, status=status.HTTP_401_UNAUTHORIZED)
    if not self.permission_classes:
        return model_class.objects.all()
    all_permissions = user.get_all_allow_perms()
    has_perm_id = []

    content = ContentType.objects.get_for_model(model_class)
    action_perm = perm_actions.get('view')
    perm_name = f'{content.app_label}_{content.model}'
    if pk:
        perm_name = f'{perm_name}_{pk}'
    print(f"PERM QUERYSET: {perm_name}")
    # Get all price list ids which required permissions
    perms_content = Perm.objects.filter(name__icontains=perm_name)
    perm_req_id = {v.object_id for v in perms_content if v.object_id}
    # print(f"PERM perm_req_id: {perm_req_id}")

    # Get all price list ids which user has permissions
    for perm in all_permissions:
        if perm.startswith(action_perm + '_' + perm_name):
            _, object_id = perm.rsplit('_', 1)
            has_perm_id.append(object_id)
    # print(f"PERM has_perm_id: {has_perm_id}")

    # Exclude item which use not has permission
    exclude_id = list(perm_req_id - set(has_perm_id))

    queryset = model_class.objects.exclude(id__in=exclude_id)
    print(f"--------------")
    print(f"Exclude id: {exclude_id}")
    # Check if the model has a 'status' field and exclude deactivated items
    if hasattr(model_class, 'status'):
        queryset = queryset.exclude(status='deactivate')
    return queryset


def get_full_permname(model, action, pk):
    perm_name = get_perm_name(model)
    if pk:
        perm_name = f'{perm_name}_{pk}'
    return f'{action}_{perm_name}'


def get_user_by_permname(perm_name):
    user_group = User.objects.filter(
        Q(group_user__perm__name=perm_name, group_user__usergroupperm__allow=True) |
        Q(perm_user__name=perm_name, perm_user__userperm__allow=True)
    ).distinct().values_list('id', flat=True)
    return list(user_group)


def export_users_has_perm(model: Model, pk: str):
    perm_name = get_perm_name(model)
    perm_name_pk = perm_name + f'_{pk}'
    print(f"Test permname: {perm_name_pk}")
    # Tìm các nhóm có quyền liên quan và không phải là admin
    group_perm_has_perm = GroupPermPerms.objects.filter(
        perm__name__iendswith=perm_name_pk
    ).values_list('group__name', flat=True).distinct()
    groups_with_perm = (GroupPerm.objects.filter(name__in=group_perm_has_perm)
                        .exclude(name='admin').distinct())
    print(f"Check group with perm: {groups_with_perm}")
    # Tìm các user có UserPerm phù hợp và không thuộc nhóm admin
    users_with_group_perm = User.objects.filter(
        usergroupperm__group__in=groups_with_perm,
        # usergroupperm__allow=True
    ).distinct()
    print(f"Check user in group has perm: {users_with_group_perm}")
    users_with_direct_perm = User.objects.filter(
        userperm__perm__name__icontains=perm_name_pk,
        # userperm__allow=True
    ).distinct()
    print(f"Check user has perm: {users_with_direct_perm}")

    # Hợp nhất hai QuerySet
    users_with_perm = users_with_group_perm.union(users_with_direct_perm)
    print(f"Check final result: {users_with_perm}")
    users_with_perm = users_with_perm.values_list('id', flat=True)
    # Xử lý kết quả, chẳng hạn tạo response hoặc log thông tin
    app_log.info(
        f"Found {users_with_perm.count()} users with permission '{perm_name_pk}' not in 'admin' group.")

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'maKH'

    for index, user_id in enumerate(users_with_perm, start=2):
        sheet[f'{get_column_letter(1)}{index}'] = user_id

    return workbook
    # # Chuẩn bị response trả về
    # response = HttpResponse(
    #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    # )
    # response['Content-Disposition'] = f'attachment; filename="UserDungBangGia_{pk}.xlsx"'
    #
    # # Lưu workbook vào response
    # workbook.save(response)
    #
    # return response
