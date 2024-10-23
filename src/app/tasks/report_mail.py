import time
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.core.mail import EmailMessage
from django.db.models import ExpressionWrapper, DurationField, F, Max, Q, Sum
from django.db.models.functions import ExtractDay
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from account.models import User
from app.logs import app_log
from marketing.order.api.views import generate_order_excel
from marketing.order.models import Order, OrderDetail
from system_func.models import PeriodSeason
from user_system.client_profile.models import ClientProfile
from user_system.daily_email.models import EmailDetail, UserGetMail
from utils.constants import mail_type
from utils.helpers import local_time


def send_daily_email(date_get, email=None):
    email_detail = EmailDetail.objects.filter(email_type=mail_type.report_order).first()

    user_get_mails = UserGetMail.objects.filter(email_detail=email_detail)

    # orders = Order.objects.filter(date_company_get=date_get).exclude(status='deactivate').order_by('-date_get')
    next_date_get = date_get + timedelta(days=1)
    orders = (Order.objects.filter(date_company_get__gte=date_get, date_company_get__lt=next_date_get)
              .exclude(status='deactivate', client_id__group_user__name='test').order_by('-date_get'))

    workbook = generate_order_excel(orders, date_get, True)
    workbook = add_new_sheet(workbook, date_get)
    workbook = add_sheet_product(workbook, orders, next_date_get)

    excel_data = BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)

    context = {
        'subject': email_detail.name,
        'body_message': email_detail.description,
    }
    # emails = ["vdt1073@gmail.com"]
    if email is not None:
        emails = email
    else:
        emails = [user_get_mail.user.email for user_get_mail in user_get_mails
                  if user_get_mail.user.email and user_get_mail.user.email != '']
    app_log.info(f"Send mail to: {emails}")
    send_bulk_email(emails, context, date_get, excel_data)

    # Cập nhật last_sent cho từng user
    for user_get_mail in user_get_mails:
        user_get_mail.last_sent = local_time()
        user_get_mail.save()


def send_bulk_email(emails, context, date_get=None, excel_data=None):
    try:
        try:
            date_get = date_get.date()
        except Exception:
            date_get = ''
        email = EmailMessage(
            subject=context['subject'],
            body=context['body_message'],
            from_email=settings.EMAIL_HOST_USER,
            to=emails,
        )
        file_name = f"BaoCaoToa_{date_get}.xlsx"

        if excel_data:
            email.attach(file_name, excel_data.read(),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()
    except Exception as e:
        app_log.error(f"got error in send daily email: \n{e}")


def send_daily_nvtt_email(date_get):
    print(f"Date get: {date_get}")
    nvtts = User.objects.filter(group_user__name='nvtt').order_by('id')

    next_date_get = date_get + timedelta(days=1)
    orders = (Order.objects.filter(date_company_get__gte=date_get, date_company_get__lt=next_date_get)
              .exclude(status='deactivate').order_by('-date_get'))
    for nvtt in nvtts:
        if nvtt.id == 'NVTTTEST':
            orders_of_nvtt = orders.filter(client_id__clientprofile__nvtt_id=nvtt.id)
            if orders_of_nvtt.count() == 0:
                context = {
                    'subject': nvtt.email,
                    'body_message': "Đại lý không có toa mới",
                }

                send_bulk_email([nvtt.email], context=context)
            else:
                workbook = generate_order_excel(orders, date_get)

                excel_data = BytesIO()
                workbook.save(excel_data)
                excel_data.seek(0)
                context = {
                    'subject': nvtt.email,
                    'body_message': "Báo cáo toa của đại lý",
                }

                send_bulk_email([nvtt.email], context, date_get, excel_data)


def get_clients():
    current_period = PeriodSeason.get_period_by_date('turnover')
    date_last_period = current_period.from_date - timedelta(days=1)

    last_period = PeriodSeason.objects.filter(from_date__lte=date_last_period,
                                              to_date__gte=date_last_period,
                                              type='turnover').first()
    current_date = datetime.now().date()

    date_til_not_orders: int = 30
    cut_off_date = current_date - timedelta(days=date_til_not_orders)

    # Lấy ra những clients có đơn hàng cách đây hơn 10 ngày
    clients_with_orders = User.objects.filter(
        order__date_get__gte=last_period.from_date,
        order__date_get__lt=cut_off_date
    ).exclude(order__date_get__gte=cut_off_date).annotate(
        latest_order_date=Max('order__date_get')
    ).annotate(
        days_diff=ExtractDay(ExpressionWrapper(
            current_date - F('latest_order_date'),
            output_field=DurationField()
        )),
        nvtt_id=F('clientprofile__nvtt_id'),
    ).order_by('nvtt_id', 'days_diff').distinct()

    return clients_with_orders


def add_new_sheet(workbook, date_get):
    clients = get_clients()
    sheet = workbook.create_sheet("KhachKoPhatSinhToa")

    # Các tiêu đề cột
    columns = ['Mã KH', 'Tên KH', 'NPP', 'NVTT', 'Ngày chưa báo toa', 'Lần cuối đặt toa']
    sheet.append(columns)

    # Định nghĩa font và alignment
    bold_font = Font(bold=True)

    # Định nghĩa độ rộng cột
    column_widths = {
        'A': 12,
        'B': 24,
        'C': 18,
        'D': 18,
        'E': 16,
        'F': 14
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    # Thêm dữ liệu và áp dụng định dạng
    for client in clients:
        try:
            npp_id = client.clientprofile.client_lv1_id
            npp = User.objects.get(id=npp_id)
            npp_name = npp.clientprofile.register_name
        except Exception:
            npp_name = ''

        try:
            nvtt_id = client.clientprofile.nvtt_id
            nvtt = User.objects.get(id=nvtt_id)
            nvtt_name = nvtt.employeeprofile.register_name
        except Exception:
            nvtt_name = ''

        client_data = [
            client.id,
            client.clientprofile.register_name if hasattr(client, 'clientprofile') else '',
            npp_name,
            nvtt_name,
            client.days_diff,
            (date_get - timedelta(days=client.days_diff)).strftime('%Y-%m-%d')
        ]
        row_idx = sheet.max_row + 1
        sheet.append(client_data)

        for col_idx in range(1, len(client_data) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            # Áp dụng font đậm cho cột 'Ngày chưa báo toa'
            if col_idx == 5:
                cell.font = bold_font

    return workbook


def format_number(value):
    if value is None:
        return Decimal(0)
    value = str(value).replace(',', '')
    return Decimal(value)


def add_sheet_product(workbook, orders, next_date_get):
    # Get data for pre-queries
    current_period = PeriodSeason.get_period_by_date('turnover')
    last_period = PeriodSeason.objects.filter(
        from_date__lte=current_period.from_date - timedelta(days=1),
        to_date__gte=current_period.from_date - timedelta(days=1),
        type='turnover'
    ).first()
    next_date_get_last = next_date_get - relativedelta(years=1)

    # Exclude items queries
    exclude_order = Q(status='deactivate', client_id__group_user__name='test',
                      order_detail__product_id_id__in=['KTUDLH'])
    # Get today products got updating
    product_ids = orders.values_list('order_detail__product_id', 'order_detail__product_id__name').distinct()

    # Query current period data orders
    orders_current_data = Order.objects.filter(
        date_company_get__gte=current_period.from_date,
        date_company_get__lt=next_date_get,
    ).exclude(exclude_order).values('npp_id', 'order_detail__product_id').annotate(
        total_boxes=Sum('order_detail__order_box')
    )

    # Query last period data orders
    orders_last_data = Order.objects.filter(
        date_company_get__gte=last_period.from_date,
        date_company_get__lt=next_date_get_last
    ).exclude(exclude_order).values('npp_id', 'order_detail__product_id').annotate(
        total_boxes=Sum('order_detail__order_box')
    )

    # Query today orders
    orders_today_data = orders.values('npp_id', 'order_detail__product_id').annotate(
        total_boxes=Sum('order_detail__order_box')
    )

    # Create excel sheet
    product_sheet = workbook.create_sheet("Báo cáo mặt hàng")
    bold_font = Font(bold=True)
    medium_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                           top=Side(style='medium'), bottom=Side(style='medium'))
    column_widths = {
        'Mã hàng': 24,
        'Tên NPP': 24,
        'Kì trước': 16,
        'Kì này': 16,
        'Phát sinh': 16
    }
    columns = list(column_widths.keys())
    product_sheet.append(columns)
    for i, column in enumerate(columns, start=1):
        column_letter = get_column_letter(i)
        product_sheet.column_dimensions[column_letter].width = column_widths[column]
        cell = product_sheet.cell(row=1, column=i)
        cell.font = bold_font
        cell.border = medium_border

    # Get npp_id list
    npp_ids_today = orders.values_list('npp_id', flat=True).distinct()
    npp_ids_last = orders_last_data.values_list('npp_id', flat=True).distinct()

    npp_ids = (npp_ids_today | npp_ids_last).distinct()

    # Get npp profile with id: name
    npp_profiles = ClientProfile.objects.filter(client_id_id__in=npp_ids).values_list('client_id_id', 'register_name')
    npp_profiles_dict = dict(npp_profiles)
    # Loop product to add the comparison of last and current period
    start_time = time.time()
    for product_id, product_name in dict(product_ids).items():
        total_last = Decimal(0)
        total_current = Decimal(0)
        total_today = Decimal(0)
        for npp_id in set(npp_ids):
            product_boxes_last = format_number(
                orders_last_data.filter(npp_id=npp_id, order_detail__product_id=product_id).aggregate(
                    Sum('total_boxes'))['total_boxes__sum'])
            product_boxes = format_number(
                orders_current_data.filter(npp_id=npp_id, order_detail__product_id=product_id).aggregate(
                    Sum('total_boxes'))['total_boxes__sum'])
            product_boxes_today = format_number(
                orders_today_data.filter(npp_id=npp_id, order_detail__product_id=product_id).aggregate(
                    Sum('total_boxes'))['total_boxes__sum'])
            if float(product_boxes_last) == 0 and float(product_boxes) == 0 and float(product_boxes_today) == 0:
                continue
            else:
                total_last += product_boxes_last
                total_current += product_boxes
                total_today += product_boxes_today

                npp_name = npp_profiles_dict.get(npp_id, '')
                product_boxes_today = "{:,.2f}".format(product_boxes_today) if float(product_boxes_today) != 0 else ''
                append_data = [
                    product_id,
                    npp_name,
                    "{:,.2f}".format(product_boxes_last),
                    "{:,.2f}".format(product_boxes),
                    product_boxes_today
                ]
                product_sheet.append(append_data)
        row = [product_name, 'Tổng cộng', "{:,.2f}".format(total_last), "{:,.2f}".format(total_current),
               "{:,.2f}".format(total_today)]
        product_sheet.append(row)
        for col in range(1, len(row) + 1):
            cell = product_sheet.cell(row=product_sheet.max_row, column=col)
            cell.font = bold_font
            cell.border = medium_border
    print(f"End looping: {time.time() - start_time}")
    return workbook
